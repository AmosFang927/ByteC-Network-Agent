#!/usr/bin/env python3
"""
Search Affiliate Orders API Excel可视化模块
将API响应数据转换为Excel格式
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Union

logger = logging.getLogger(__name__)

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("⚠️ Excel功能不可用，请安装 pandas 和 openpyxl: pip install pandas openpyxl")

class ExcelVisualizer:
    """Excel可视化类"""
    
    def __init__(self, output_dir: str = "output"):
        """
        初始化Excel可视化器
        
        Args:
            output_dir: 输出目录
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("Excel功能不可用，请安装 pandas 和 openpyxl")
        
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        logger.info(f"🔧 ExcelVisualizer 初始化完成，输出目录: {self.output_dir}")
    
    def _format_timestamp(self, timestamp: Optional[int]) -> str:
        """
        格式化时间戳
        
        Args:
            timestamp: Unix时间戳
            
        Returns:
            格式化的时间字符串
        """
        if timestamp is None:
            return "N/A"
        
        try:
            dt = datetime.fromtimestamp(timestamp)
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except (ValueError, OSError):
            return f"Invalid timestamp: {timestamp}"
    
    def _format_amount(self, amount: Optional[Union[int, float]]) -> str:
        """
        格式化金额
        
        Args:
            amount: 金额数值
            
        Returns:
            格式化的金额字符串
        """
        if amount is None:
            return "0.00"
        
        try:
            return f"{float(amount):.2f}"
        except (ValueError, TypeError):
            return "0.00"
    
    def _flatten_order_data(self, order: Dict[str, Any]) -> Dict[str, Any]:
        """
        扁平化订单数据
        
        Args:
            order: 原始订单数据
            
        Returns:
            扁平化的订单数据
        """
        flattened = {}
        
        # 基本信息
        flattened['订单ID'] = order.get('orderId', 'N/A')
        flattened['订单状态'] = order.get('status', 'N/A')
        flattened['创建时间'] = self._format_timestamp(order.get('createTime'))
        flattened['更新时间'] = self._format_timestamp(order.get('updateTime'))
        
        # 金额信息
        flattened['订单金额'] = self._format_amount(order.get('orderAmount'))
        flattened['佣金金额'] = self._format_amount(order.get('commissionAmount'))
        flattened['佣金率'] = f"{order.get('commissionRate', 0):.2f}%"
        
        # 产品信息
        product = order.get('product', {})
        flattened['产品ID'] = product.get('productId', 'N/A')
        flattened['产品名称'] = product.get('productName', 'N/A')
        flattened['产品价格'] = self._format_amount(product.get('price'))
        flattened['产品数量'] = product.get('quantity', 0)
        
        # 活动信息
        campaign = order.get('campaign', {})
        flattened['活动ID'] = campaign.get('campaignId', 'N/A')
        flattened['活动名称'] = campaign.get('campaignName', 'N/A')
        
        # 用户信息
        user = order.get('user', {})
        flattened['用户ID'] = user.get('userId', 'N/A')
        flattened['用户名'] = user.get('userName', 'N/A')
        
        # 其他信息
        flattened['订单来源'] = order.get('orderSource', 'N/A')
        flattened['支付方式'] = order.get('paymentMethod', 'N/A')
        flattened['备注'] = order.get('remark', '')
        
        return flattened
    
    def _create_summary_data(self, orders: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        创建汇总数据
        
        Args:
            orders: 订单列表
            
        Returns:
            汇总数据字典
        """
        if not orders:
            return {
                '总订单数': 0,
                '总金额': '0.00',
                '总佣金': '0.00',
                '平均佣金率': '0.00%',
                '状态分布': {},
                '活动分布': {},
                '时间分布': {}
            }
        
        # 基础统计
        total_orders = len(orders)
        total_amount = sum(float(order.get('orderAmount', 0)) for order in orders)
        total_commission = sum(float(order.get('commissionAmount', 0)) for order in orders)
        avg_commission_rate = (total_commission / total_amount * 100) if total_amount > 0 else 0
        
        # 状态分布
        status_distribution = {}
        for order in orders:
            status = order.get('status', 'Unknown')
            status_distribution[status] = status_distribution.get(status, 0) + 1
        
        # 活动分布
        campaign_distribution = {}
        for order in orders:
            campaign_name = order.get('campaign', {}).get('campaignName', 'Unknown')
            campaign_distribution[campaign_name] = campaign_distribution.get(campaign_name, 0) + 1
        
        # 时间分布（按天）
        time_distribution = {}
        for order in orders:
            create_time = order.get('createTime')
            if create_time:
                try:
                    date_str = datetime.fromtimestamp(create_time).strftime("%Y-%m-%d")
                    time_distribution[date_str] = time_distribution.get(date_str, 0) + 1
                except (ValueError, OSError):
                    pass
        
        return {
            '总订单数': total_orders,
            '总金额': f"{total_amount:.2f}",
            '总佣金': f"{total_commission:.2f}",
            '平均佣金率': f"{avg_commission_rate:.2f}%",
            '状态分布': status_distribution,
            '活动分布': campaign_distribution,
            '时间分布': time_distribution
        }
    
    def export_to_excel(self, api_response: Dict[str, Any], filename: Optional[str] = None) -> str:
        """
        导出数据到Excel文件
        
        Args:
            api_response: API响应数据
            filename: 文件名（可选）
            
        Returns:
            输出文件路径
        """
        if not api_response.get("success"):
            logger.error(f"❌ API响应失败，无法导出数据: {api_response.get('error')}")
            raise ValueError(f"API响应失败: {api_response.get('error')}")
        
        orders = api_response.get("orders", [])
        if not orders:
            logger.warning("⚠️ 没有订单数据可导出")
            return ""
        
        # 生成文件名
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"affiliate_orders_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        
        # 创建Excel工作簿
        wb = Workbook()
        
        # 删除默认工作表
        wb.remove(wb.active)
        
        # 创建订单详情工作表
        ws_orders = wb.create_sheet("订单详情")
        
        # 扁平化数据
        flattened_orders = [self._flatten_order_data(order) for order in orders]
        
        # 创建DataFrame
        df = pd.DataFrame(flattened_orders)
        
        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws_orders.append(r)
        
        # 设置样式
        self._format_excel_worksheet(ws_orders)
        
        # 创建汇总工作表
        ws_summary = wb.create_sheet("数据汇总")
        summary_data = self._create_summary_data(orders)
        self._create_summary_worksheet(ws_summary, summary_data)
        
        # 创建图表工作表
        ws_charts = wb.create_sheet("数据图表")
        self._create_charts_worksheet(ws_charts, summary_data)
        
        # 保存文件
        wb.save(filepath)
        
        logger.info(f"✅ Excel文件导出成功: {filepath}")
        logger.info(f"📊 导出了 {len(orders)} 条订单记录")
        
        return str(filepath)
    
    def _format_excel_worksheet(self, ws):
        """
        格式化Excel工作表
        
        Args:
            ws: 工作表对象
        """
        # 设置列宽
        column_widths = {
            'A': 20,  # 订单ID
            'B': 15,  # 订单状态
            'C': 20,  # 创建时间
            'D': 20,  # 更新时间
            'E': 15,  # 订单金额
            'F': 15,  # 佣金金额
            'G': 12,  # 佣金率
            'H': 20,  # 产品ID
            'I': 30,  # 产品名称
            'J': 15,  # 产品价格
            'K': 12,  # 产品数量
            'L': 20,  # 活动ID
            'M': 25,  # 活动名称
            'N': 20,  # 用户ID
            'O': 20,  # 用户名
            'P': 15,  # 订单来源
            'Q': 15,  # 支付方式
            'R': 30,  # 备注
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 设置数据行样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
    
    def _create_summary_worksheet(self, ws, summary_data: Dict[str, Any]):
        """
        创建汇总工作表
        
        Args:
            ws: 工作表对象
            summary_data: 汇总数据
        """
        # 基础统计
        ws['A1'] = "基础统计"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = "总订单数"
        ws['B3'] = summary_data['总订单数']
        
        ws['A4'] = "总金额"
        ws['B4'] = summary_data['总金额']
        
        ws['A5'] = "总佣金"
        ws['B5'] = summary_data['总佣金']
        
        ws['A6'] = "平均佣金率"
        ws['B6'] = summary_data['平均佣金率']
        
        # 状态分布
        ws['D1'] = "订单状态分布"
        ws['D1'].font = Font(bold=True, size=14)
        
        row = 3
        for status, count in summary_data['状态分布'].items():
            ws[f'D{row}'] = status
            ws[f'E{row}'] = count
            row += 1
        
        # 活动分布
        ws['G1'] = "活动分布"
        ws['G1'].font = Font(bold=True, size=14)
        
        row = 3
        for campaign, count in summary_data['活动分布'].items():
            ws[f'G{row}'] = campaign
            ws[f'H{row}'] = count
            row += 1
        
        # 时间分布
        ws['J1'] = "时间分布"
        ws['J1'].font = Font(bold=True, size=14)
        
        row = 3
        for date, count in summary_data['时间分布'].items():
            ws[f'J{row}'] = date
            ws[f'K{row}'] = count
            row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 10
    
    def _create_charts_worksheet(self, ws, summary_data: Dict[str, Any]):
        """
        创建图表工作表
        
        Args:
            ws: 工作表对象
            summary_data: 汇总数据
        """
        ws['A1'] = "数据图表"
        ws['A1'].font = Font(bold=True, size=16)
        
        # 状态分布图表数据
        ws['A3'] = "状态分布"
        ws['A3'].font = Font(bold=True, size=12)
        
        row = 5
        ws[f'A{row}'] = "状态"
        ws[f'B{row}'] = "数量"
        row += 1
        
        for status, count in summary_data['状态分布'].items():
            ws[f'A{row}'] = status
            ws[f'B{row}'] = count
            row += 1
        
        # 活动分布图表数据
        ws['D3'] = "活动分布"
        ws['D3'].font = Font(bold=True, size=12)
        
        row = 5
        ws[f'D{row}'] = "活动"
        ws[f'E{row}'] = "数量"
        row += 1
        
        for campaign, count in summary_data['活动分布'].items():
            ws[f'D{row}'] = campaign
            ws[f'E{row}'] = count
            row += 1
        
        # 时间分布图表数据
        ws['G3'] = "时间分布"
        ws['G3'].font = Font(bold=True, size=12)
        
        row = 5
        ws[f'G{row}'] = "日期"
        ws[f'H{row}'] = "数量"
        row += 1
        
        for date, count in summary_data['时间分布'].items():
            ws[f'G{row}'] = date
            ws[f'H{row}'] = count
            row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 10

# 便捷函数
def export_orders_to_excel(api_response: Dict[str, Any], output_dir: str = "output") -> str:
    """
    导出订单数据到Excel文件
    
    Args:
        api_response: API响应数据
        output_dir: 输出目录
        
    Returns:
        输出文件路径
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("Excel功能不可用，请安装 pandas 和 openpyxl")
    
    visualizer = ExcelVisualizer(output_dir)
    return visualizer.export_to_excel(api_response) 