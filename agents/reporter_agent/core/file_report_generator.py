#!/usr/bin/env python3
"""
文件報告生成器
專門處理從DMP Agent輸出文件中讀取的數據並生成報告
支持完整的Partner ALL邏輯、多sheets、統計分析等
"""

import os
import sys
import re
import pandas as pd
import time
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
import logging
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from dataclasses import dataclass
from decimal import Decimal

# 添加項目根目錄到路徑
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))))

import config
from modules.feishu_uploader import FeishuUploader
from agents.data_output_agent.email_sender import EmailSender
from utils.logger import print_step

logger = logging.getLogger(__name__)

@dataclass
class PartnerSummary:
    """Partner匯總數據類 - 與原有ReportGenerator相容"""
    partner_name: str
    partner_id: Optional[int]
    total_records: int
    total_amount: Decimal
    sources: List[str]
    excluded_records: int = 0
    excluded_statuses: List[str] = None
    file_path: Optional[str] = None
    
    def __post_init__(self):
        if self.excluded_statuses is None:
            self.excluded_statuses = []
    
    @property
    def amount_formatted(self) -> str:
        return f"${self.total_amount:,.2f}"
    
    @property
    def sources_count(self) -> int:
        return len(self.sources)
    
    @property
    def has_excluded_records(self) -> bool:
        return self.excluded_records > 0
    
    def to_dict(self) -> Dict[str, Any]:
        """轉換為字典格式"""
        return {
            'partner_name': self.partner_name,
            'partner_id': self.partner_id,
            'total_records': self.total_records,
            'total_amount': float(self.total_amount) if self.total_amount else 0.0,
            'amount_formatted': self.amount_formatted,
            'sources': self.sources if self.sources else [],
            'sources_count': self.sources_count,
            'excluded_records': self.excluded_records,
            'excluded_statuses': self.excluded_statuses
        }


class FileReportGenerator:
    """文件報告生成器 - 處理DMP Agent輸出的文件數據"""
    
    def __init__(self, output_dir: str = "output", global_email_disabled: bool = False):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # 初始化飞书上传和邮件发送器
        self.feishu_uploader = FeishuUploader()
        self.email_sender = EmailSender(global_email_disabled=global_email_disabled)
    
    async def generate_report_from_file(self, df: pd.DataFrame, 
                                      partner_name: str,
                                      import_file_path: str,
                                      send_email: bool = True,
                                      upload_feishu: bool = True,
                                      self_email: bool = False,
                                      start_date: Optional[str] = None,
                                      end_date: Optional[str] = None) -> Dict[str, Any]:
        """
        從DMP文件數據生成報告 - 支持完整的Partner ALL邏輯
        
        Args:
            df: DMP Agent輸出的DataFrame
            partner_name: Partner名稱
            import_file_path: 原始導入文件路徑
            send_email: 是否發送郵件
            upload_feishu: 是否上傳飛書
            self_email: 是否發送給自己
            
        Returns:
            報告生成結果
        """
        try:
            logger.info(f"🚀 開始從文件數據生成 {partner_name} 報告")
            # logger.info(f"📊 數據概況: {len(df):,} 行, {len(df.columns)} 列")
            
            # 1. 標準化數據格式並添加Partner分組
            standardized_df = await self._standardize_and_classify_data(df)
            
            # 2. 根據partner_name處理數據
            if partner_name.upper() == 'ALL':
                return await self._generate_all_partners_report(
                    standardized_df, import_file_path, send_email, upload_feishu, self_email, start_date, end_date
                )
            else:
                return await self._generate_single_partner_report(
                    standardized_df, partner_name, import_file_path, send_email, upload_feishu, self_email, start_date, end_date
                )
                
        except Exception as e:
            logger.error(f"❌ 文件報告生成失敗: {e}")
            return {
                'success': False,
                'error': str(e),
                'partner_name': partner_name,
                'total_records': 0,
                'total_amount': 0.0,
                'excel_files': []
            }
    
    async def _standardize_and_classify_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """DMP Agent已經標準化數據，這裡只需要基本驗證"""
        try:
            # logger.info("🔄 驗證DMP Agent標準化數據...")
            
            # 導入配置以驗證標準欄位
            standard_columns = config.get_standard_report_columns()
            
            # 檢查數據是否包含必要的標準欄位
            missing_columns = []
            # 檢查多種可能的欄位名稱格式
            amount_column = None
            possible_amount_columns = ['USD Sale Amount', 'Sale Amount (USD)', 'Sale Amount', 'usd_sale_amount']
            for col in possible_amount_columns:
                if col in df.columns:
                    amount_column = col
                    break
            
            for col in ['Partner', 'Source', amount_column, 'Status']:
                if col is None:
                    continue
                if col not in df.columns:
                    missing_columns.append(col)
            
            if missing_columns:
                # logger.warning(f"⚠️ 缺少關鍵欄位: {missing_columns}")
                # logger.info("🔄 嘗試基本修復...")
                
                # 基本修復邏輯
                if 'Partner' not in df.columns:
                    df['Partner'] = 'Unknown'
                if 'Source' not in df.columns:
                    df['Source'] = df.get('Aff Sub1', df.get('Aff Sub', 'Unknown'))
                # 確保有金額欄位
                if amount_column is None:
                    df['USD Sale Amount'] = 0.0
                    amount_column = 'USD Sale Amount'
                if 'Status' not in df.columns:
                    df['Status'] = 'Pending'
            
            # 添加欄位名稱標準化映射
            column_mapping = {
                'Conversion Date': 'Datetime Conversion',
                'Sale Amount (USD)': 'USD Sale Amount',
                'Publisher Sub ID 1': 'Aff Sub1',
                'Publisher Sub ID 2': 'Aff Sub2', 
                'Publisher Sub ID 3': 'Aff Sub3',
                'Publisher Sub ID 4': 'Aff Sub4',
                'Publisher Sub ID 5': 'Aff Sub5',
                'Advertiser Sub ID 2': 'Adv Pub2',
                'Advertiser Sub ID 3': 'Adv Pub3',
                'Advertiser Sub ID 4': 'Adv Pub4', 
                'Advertiser Sub ID 5': 'Adv Pub5'
            }
            
            # 執行欄位重新命名
            df = df.rename(columns=column_mapping)
            logger.info(f"✅ 欄位標準化完成，更新後欄位: {list(df.columns)}")
            
            # 更新 amount_column 的引用
            if amount_column == 'Sale Amount (USD)':
                amount_column = 'USD Sale Amount'
            
            # 🚨 重要：應用 DMP_PASSTHROUGH_REMOVE_COLUMNS 移除不需要的欄位
            try:
                columns_to_remove = config.DMP_PASSTHROUGH_REMOVE_COLUMNS
                existing_columns_to_remove = [col for col in columns_to_remove if col in df.columns]
                
                if existing_columns_to_remove:
                    df = df.drop(columns=existing_columns_to_remove)
                    logger.info(f"✅ 已移除 DMP Passthrough 欄位: {existing_columns_to_remove}")
                else:
                    logger.info("📋 無需移除的 DMP Passthrough 欄位")
                    
            except Exception as e:
                logger.warning(f"⚠️ 應用 DMP_PASSTHROUGH_REMOVE_COLUMNS 時出錯: {e}")
            
            # 修復 Datetime Conversion 欄位（如果為空）
            if 'Datetime Conversion' in df.columns:
                missing_dates = df['Datetime Conversion'].isna().sum()
                if missing_dates > 0:
                    logger.warning(f"⚠️ 發現 {missing_dates} 個空的 Datetime Conversion，使用默認日期修復")
                    # 嘗試從文件名提取日期，或使用昨天日期
                    default_date_str = self._extract_date_from_filename(import_file_path)
                    if default_date_str:
                        default_date = datetime.strptime(default_date_str, '%Y-%m-%d')
                    else:
                        default_date = datetime.now() - timedelta(days=1)
                    df['Datetime Conversion'] = df['Datetime Conversion'].fillna(default_date)
                    logger.info(f"✅ 已使用 {default_date.strftime('%Y-%m-%d')} 填充 {missing_dates} 個空的日期欄位")
            
            # 確保數值類型正確
            numeric_columns = [amount_column, 'USD Payout', 'Conversion ID', 'Offer ID']
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            
            # logger.info(f"✅ 數據驗證完成，包含列: {list(df.columns)}")
            # logger.info(f"📊 Partner分布: {df['Partner'].value_counts().to_dict()}")
            
            return df
            
        except Exception as e:
            logger.error(f"❌ 數據驗證失敗: {e}")
            return df
    

    
    async def _generate_all_partners_report(self, df: pd.DataFrame, import_file_path: str,
                                          send_email: bool, upload_feishu: bool, self_email: bool,
                                          start_date: Optional[str] = None, end_date: Optional[str] = None) -> Dict[str, Any]:
        """生成所有Partner的報告"""
        try:
            # logger.info("🌐 生成所有Partner報告...")
            
            # 1. 生成Partner統計信息
            partner_summaries = await self._create_partner_summaries(df)
            
            # 2. 為每個Partner生成獨立的Excel文件
            excel_files = []
            all_total_records = 0
            all_total_amount = 0.0
            
            for summary in partner_summaries:
                if summary.total_records == 0:
                    # logger.info(f"⚠️ {summary.partner_name} 沒有數據，跳過Excel生成")
                    continue
                    
                # 過濾該Partner的數據
                partner_df = df[df['Partner'] == summary.partner_name].copy()
                
                # 生成該Partner的Excel文件
                excel_file = await self._create_partner_excel_file(
                    partner_df, summary, import_file_path, start_date, end_date
                )
                
                if excel_file:
                    excel_files.append(excel_file)
                    summary.file_path = excel_file
                    all_total_records += summary.total_records
                    all_total_amount += float(summary.total_amount)
                    
                    logger.info(f"📄 {summary.partner_name}: {summary.total_records:,} 記錄, ${summary.total_amount:,.2f}")
            
            # 3. 創建結果
            result = {
                'success': True,
                'partner_name': 'ALL',
                'total_records': all_total_records,
                'total_amount': all_total_amount,
                'excel_files': excel_files,
                'import_file': import_file_path,
                'partner_summaries': [s.to_dict() for s in partner_summaries],
                'generation_time': datetime.now().isoformat()
            }
            
            # 4. 發送郵件
            if send_email and excel_files:
                await self._send_all_partner_emails(excel_files, self_email, start_date, end_date)
            
            # 5. 上傳飛書
            if upload_feishu and excel_files:
                await self._upload_all_partner_feishu(excel_files)
            
            logger.info(f"✅ 所有Partner報告生成完成: {len(excel_files)} 個文件")
            return result
            
        except Exception as e:
            logger.error(f"❌ 生成所有Partner報告失敗: {e}")
            raise
    
    async def _generate_single_partner_report(self, df: pd.DataFrame, partner_name: str, 
                                            import_file_path: str, send_email: bool, 
                                            upload_feishu: bool, self_email: bool,
                                            start_date: Optional[str] = None, end_date: Optional[str] = None) -> Dict[str, Any]:
        """生成單個Partner的報告"""
        try:
            logger.info(f"🎯 生成 {partner_name} Partner報告...")
            
            # 過濾該Partner的數據
            partner_df = df[df['Partner'] == partner_name].copy()
            
            if partner_df.empty:
                logger.warning(f"⚠️ 沒有找到 {partner_name} 的數據")
                return {
                    'success': True,
                    'partner_name': partner_name,
                    'total_records': 0,
                    'total_amount': 0.0,
                    'excel_files': [],
                    'import_file': import_file_path
                }
            
            # 創建Partner統計
            partner_summary = await self._create_single_partner_summary(partner_df, partner_name)
            
            # 生成Excel文件
            excel_file = await self._create_partner_excel_file(
                partner_df, partner_summary, import_file_path, start_date, end_date
            )
            
            excel_files = [excel_file] if excel_file else []
            
            # 設置Partner Summary的file_path
            if excel_file:
                partner_summary.file_path = excel_file
                logger.info(f"✅ 設置 {partner_name} 的file_path: {excel_file}")
            else:
                logger.warning(f"⚠️ {partner_name} 的excel_file為None")
            
            # 創建結果
            result = {
                'success': True,
                'partner_name': partner_name,
                'total_records': partner_summary.total_records,
                'total_amount': float(partner_summary.total_amount),
                'excel_files': excel_files,
                'import_file': import_file_path,
                'generation_time': datetime.now().isoformat()
            }
            
            # 發送郵件和上傳飛書
            if send_email and excel_files:
                await self._send_single_partner_email(partner_summary, self_email)
            
            if upload_feishu and excel_files:
                await self._upload_single_partner_feishu(excel_files, partner_name)
            
            # 生成單個Partner的總結（只包含當前Partner）
            await self._print_final_summary(partner_df, [partner_summary], import_file_path)
            
            return result
            
        except Exception as e:
            logger.error(f"❌ 生成 {partner_name} Partner報告失敗: {e}")
            raise
    
    async def _create_partner_summaries(self, df: pd.DataFrame) -> List[PartnerSummary]:
        """創建所有Partner的統計信息"""
        try:
            partner_summaries = []
            
            # 導入配置以獲取所有Partners
            from config import PARTNER_SOURCES_MAPPING
            
            # 獲取配置中的所有Partners
            config_partners = set(PARTNER_SOURCES_MAPPING.keys())
            
            # 按Partner分組統計（基於數據中實際存在的Partners）
            partner_groups = df.groupby('Partner')
            data_partners = set(partner_groups.groups.keys())
            
            # 合併配置中的Partners和數據中的Partners
            all_partners = config_partners.union(data_partners)
            
            logger.info(f"🔍 配置中的Partners: {sorted(config_partners)}")
            logger.info(f"🔍 數據中的Partners: {sorted(data_partners)}")
            logger.info(f"🔍 將處理的所有Partners: {sorted(all_partners)}")
            
            # 為所有Partners創建統計
            for partner_name in all_partners:
                if partner_name in data_partners:
                    # 數據中存在該Partner
                    group_df = partner_groups.get_group(partner_name)
                    summary = await self._create_single_partner_summary(group_df, partner_name)
                else:
                    # 數據中不存在該Partner，創建空的統計
                    logger.info(f"📊 為 {partner_name} 創建空統計（數據中無記錄）")
                    # 創建一個包含必要列的空的DataFrame，避免列缺失錯誤
                    empty_df = pd.DataFrame(columns=['Status', 'USD Sale Amount', 'Source', 'Partner'])
                    summary = await self._create_single_partner_summary(empty_df, partner_name)
                
                partner_summaries.append(summary)
            
            # 按total_records排序（降序）
            partner_summaries.sort(key=lambda x: x.total_records, reverse=True)
            
            return partner_summaries
            
        except Exception as e:
            logger.error(f"❌ 創建Partner統計失敗: {e}")
            return []
    
    async def _create_single_partner_summary(self, partner_df: pd.DataFrame, partner_name: str) -> PartnerSummary:
        """創建單個Partner的統計信息"""
        try:
            # 檢查必要的列是否存在
            required_columns = ['Status', 'USD Sale Amount', 'Source']
            missing_columns = [col for col in required_columns if col not in partner_df.columns]
            
            if missing_columns:
                logger.warning(f"⚠️ {partner_name} 數據缺少必要列: {missing_columns}")
                # 如果缺少Status列，假設所有記錄都是有效的
                if 'Status' not in partner_df.columns:
                    logger.info(f"📊 {partner_name} 無Status列，假設所有記錄為有效轉化")
                    total_records = len(partner_df)
                    # 自動檢測金額欄位
                    amount_col = None
                    for col in ['USD Sale Amount', 'Sale Amount (USD)', 'Sale Amount', 'usd_sale_amount']:
                        if col in partner_df.columns:
                            amount_col = col
                            break
                    total_amount = Decimal(str(partner_df[amount_col].sum())) if amount_col else Decimal('0')
                    sources = partner_df['Source'].unique().tolist() if 'Source' in partner_df.columns else []
                    sources = [s for s in sources if pd.notna(s) and s != '']
                    
                    return PartnerSummary(
                        partner_name=partner_name,
                        partner_id=None,
                        total_records=total_records,
                        total_amount=total_amount,
                        sources=sources,
                        excluded_records=0,
                        excluded_statuses=[]
                    )
                else:
                    # 如果缺少其他列，返回空統計
                    logger.error(f"❌ {partner_name} 缺少關鍵列: {missing_columns}")
                    return PartnerSummary(
                        partner_name=partner_name,
                        partner_id=None,
                        total_records=0,
                        total_amount=Decimal('0'),
                        sources=[],
                        excluded_records=0,
                        excluded_statuses=[]
                    )
            
            # 計算有效轉化（pending/approved）
            valid_statuses = ['pending', 'approved', 'approved_pending']
            valid_mask = partner_df['Status'].str.lower().isin(valid_statuses)
            valid_df = partner_df[valid_mask]
            
            total_records = len(valid_df)
            # 支持多種金額欄位名稱格式
            amount_column = None
            possible_amount_columns = ['USD Sale Amount', 'Sale Amount (USD)', 'Sale Amount', 'usd_sale_amount']
            for col in possible_amount_columns:
                if col in partner_df.columns:
                    amount_column = col
                    break
            
            if amount_column:
                total_amount = Decimal(str(valid_df[amount_column].sum()))
            else:
                logger.warning(f"⚠️ 未找到金額欄位，可用欄位: {list(partner_df.columns)}")
                total_amount = Decimal('0')
            
            # 獲取該Partner的所有Sources
            sources = partner_df['Source'].unique().tolist()
            sources = [s for s in sources if pd.notna(s) and s != '']
            
            # 計算排除的記錄（invalid/rejected）
            invalid_statuses = ['invalid', 'rejected']
            invalid_mask = partner_df['Status'].str.lower().isin(invalid_statuses)
            excluded_records = invalid_mask.sum()
            excluded_statuses = partner_df[invalid_mask]['Status'].tolist() if excluded_records > 0 else []
            
            return PartnerSummary(
                partner_name=partner_name,
                partner_id=None,
                total_records=total_records,
                total_amount=total_amount,
                sources=sources,
                excluded_records=excluded_records,
                excluded_statuses=excluded_statuses
            )
            
        except Exception as e:
            logger.error(f"❌ 創建 {partner_name} 統計失敗: {e}")
            logger.error(f"   數據列: {list(partner_df.columns) if not partner_df.empty else '空DataFrame'}")
            logger.error(f"   數據形狀: {partner_df.shape}")
            return PartnerSummary(
                partner_name=partner_name,
                partner_id=None,
                total_records=0,
                total_amount=Decimal('0'),
                sources=[],
                excluded_records=0,
                excluded_statuses=[]
            )
    
    async def _create_partner_excel_file(self, partner_df: pd.DataFrame, 
                                       partner_summary: PartnerSummary,
                                       import_file_path: str,
                                       start_date: Optional[str] = None,
                                       end_date: Optional[str] = None) -> Optional[str]:
        """為單個Partner創建Excel文件（包含多個sheets）"""
        try:
            if partner_df.empty:
                return None
                
            logger.info(f"📊 開始生成 {partner_summary.partner_name} Excel文件...")
            
            # 生成文件名 - 使用 start_date 和 end_date
            clean_partner_name = self._clean_sheet_name(partner_summary.partner_name)
            
            # 獲取日期範圍
            date_range = self._get_date_range_for_filename(start_date, end_date, import_file_path)
            filename = f"{clean_partner_name}_ConversionReport_{date_range}.xlsx"
            filepath = self.output_dir / filename
            
            # 創建工作簿
            wb = Workbook()
            wb.remove(wb.active)  # 刪除默認sheet
            
            # 樣式設置
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            data_alignment = Alignment(horizontal="left", vertical="center")
            number_alignment = Alignment(horizontal="right", vertical="center")
            
            # 1. 創建主Summary Sheet
            summary_name = clean_partner_name[:31]
            if summary_name in wb.sheetnames:
                for i in range(1, 100):
                    candidate = f"{summary_name[:28]}_{i}"
                    if candidate not in wb.sheetnames:
                        summary_name = candidate
                        break
            summary_ws = wb.create_sheet(summary_name, 0)
            
            # 添加匯總信息
            current_row = self._add_summary_header(
                summary_ws, partner_summary.partner_name, partner_df, 
                import_file_path, partner_summary, start_date, end_date
            )
            
            # 寫入主表數據（所有該Partner的數據）
            self._write_data_to_sheet(summary_ws, partner_df, current_row, header_font, 
                                    header_fill, header_alignment, data_alignment, number_alignment)
            
            # 2. 按Source創建各個Sheet
            if 'Source' in partner_df.columns:
                sources = partner_df['Source'].unique()
                
                for source in sources:
                    if pd.isna(source) or source == '':
                        continue
                        
                    # 清理sheet名稱
                    sheet_name = self._clean_sheet_name(str(source))[:31]
                    if sheet_name in wb.sheetnames:
                        for i in range(1, 100):
                            candidate = f"{sheet_name[:28]}_{i}"
                            if candidate not in wb.sheetnames:
                                sheet_name = candidate
                                break
                    
                    # 過濾該Source的數據
                    source_df = partner_df[partner_df['Source'] == source].copy()
                    
                    if not source_df.empty:
                        ws = wb.create_sheet(sheet_name)
                        
                        # 添加Source匯總信息
                        current_row = self._add_source_summary_header(
                            ws, partner_summary.partner_name, source, source_df, import_file_path, start_date, end_date
                        )
                        
                        # 寫入Source數據
                        self._write_data_to_sheet(ws, source_df, current_row, header_font,
                                                header_fill, header_alignment, data_alignment, number_alignment)
            
            # 3. 自動調整所有sheet的列寬
            for ws in wb.worksheets:
                self._auto_adjust_columns(ws)
            
            # 保存文件
            wb.save(filepath)
            logger.info(f"✅ {partner_summary.partner_name} Excel報告已生成: {filepath}")
            
            return str(filepath)
            
        except Exception as e:
            logger.error(f"❌ 生成 {partner_summary.partner_name} Excel文件失敗: {e}")
            return None
    
    def _clean_sheet_name(self, name: str) -> str:
        """清理Excel sheet名稱"""
        # 移除不允許的字符
        import re
        cleaned = re.sub(r'[/\\*\[\]:?]', '_', str(name))
        # 限制長度
        if len(cleaned) > 31:
            cleaned = cleaned[:31]
        return cleaned
    
    def _add_summary_header(self, ws, partner_name: str, df: pd.DataFrame, 
                           import_file_path: str, partner_summary: PartnerSummary,
                           start_date: Optional[str] = None, end_date: Optional[str] = None) -> int:
        """添加Partner匯總信息到Excel sheet"""
        current_row = 1
        
        # 樣式
        title_font = Font(bold=True, size=14, color="FFFFFF")
        title_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
        info_font = Font(size=10, color="333333")
        success_font = Font(bold=True, color="008000")
        
        # 標題
        ws.cell(row=current_row, column=1, value=f"📊 {partner_name} Report Summary").font = title_font
        ws.cell(row=current_row, column=1).fill = title_fill
        current_row += 1
        
        # 空行
        current_row += 1
        
        # 日期範圍
        if start_date and end_date:
            date_range = f"{start_date} to {end_date}"
        else:
            date_range = self._get_date_range_for_filename(None, None, import_file_path) # Use the new function
        ws.cell(row=current_row, column=1, value=f"Date Range: {date_range}").font = info_font
        current_row += 1
        
        # 生成時間
        ws.cell(row=current_row, column=1, value=f"Generated Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = info_font
        current_row += 1
        
        # 統計信息
        total_conversions = len(df)
        ws.cell(row=current_row, column=1, value=f"Total Conversions (All Status): {total_conversions:,}").font = info_font
        current_row += 1
        
        if 'Status' in df.columns:
            # 有效轉化統計
            valid_mask = df['Status'].str.lower().isin(['pending', 'approved', 'approved_pending'])
            valid_count = valid_mask.sum()
            valid_amount = df[valid_mask]['USD Sale Amount'].sum()
            
            ws.cell(row=current_row, column=1, value=f"✅ Total Conversions (Pending/Approved): {valid_count:,}").font = success_font
            current_row += 1
            ws.cell(row=current_row, column=1, value=f"✅ Total Sale Amount (USD) (Pending/Approved): ${valid_amount:,.2f}").font = success_font
            current_row += 1
            
            # 無效轉化統計 - 總是顯示，即使為0
            invalid_mask = df['Status'].str.lower().isin(['invalid', 'rejected'])
            invalid_count = invalid_mask.sum()
            invalid_amount = df[invalid_mask]['USD Sale Amount'].sum() if invalid_count > 0 else 0.0
            
            ws.cell(row=current_row, column=1, value=f"⚠️ Total Conversions (Invalid/Rejected): {invalid_count:,}").font = Font(size=10, color="FF0000") # 紅色
            current_row += 1
            ws.cell(row=current_row, column=1, value=f"⚠️ Total Sale Amount (USD) (Invalid/Rejected): ${invalid_amount:,.2f}").font = Font(size=10, color="FF0000") # 紅色
            current_row += 1
        
        # Sources信息 - 顯示所有 Sources，與郵件格式一致
        if partner_summary.sources:
            sources_text = ", ".join(partner_summary.sources)  # 顯示所有 Sources
            ws.cell(row=current_row, column=1, value=f"Sources: {sources_text}").font = info_font
            current_row += 1
        
        current_row += 1  # 空一行
        return current_row
    
    def _add_source_summary_header(self, ws, partner_name: str, source: str, 
                                   df: pd.DataFrame, import_file_path: str,
                                   start_date: Optional[str] = None, end_date: Optional[str] = None) -> int:
        """添加Source匯總信息到Excel sheet"""
        current_row = 1 # Keep this line
        
        # 樣式
        title_font = Font(bold=True, size=12, color="FFFFFF")
        title_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
        info_font = Font(size=10, color="333333")
        success_font = Font(bold=True, color="008000")
        warning_font = Font(size=10, color="FF8C00")  # 橙色
        
        # 標題
        ws.cell(row=current_row, column=1, value=f"📊 {source} Report Summary").font = title_font
        ws.cell(row=current_row, column=1).fill = title_fill
        current_row += 1
        
        # 空行
        current_row += 1
        
        # 日期範圍
        if start_date and end_date:
            date_range = f"{start_date} to {end_date}"
        else:
            date_range = self._get_date_range_for_filename(None, None, import_file_path) # Use the new function
        ws.cell(row=current_row, column=1, value=f"Date Range: {date_range}").font = info_font
        current_row += 1
        
        # 生成時間
        ws.cell(row=current_row, column=1, value=f"Generated Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = info_font
        current_row += 1
        
        # 統計信息
        total_conversions = len(df)
        ws.cell(row=current_row, column=1, value=f"Total Conversions (All Status): {total_conversions:,}").font = info_font
        current_row += 1
        
        if 'Status' in df.columns:
            # 有效轉化統計
            valid_mask = df['Status'].str.lower().isin(['pending', 'approved', 'approved_pending'])
            valid_count = valid_mask.sum()
            valid_amount = df[valid_mask]['USD Sale Amount'].sum()
            
            ws.cell(row=current_row, column=1, value=f"✅ Total Conversions (Pending/Approved): {valid_count:,}").font = success_font
            current_row += 1
            ws.cell(row=current_row, column=1, value=f"✅ Total Sale Amount (USD) (Pending/Approved): ${valid_amount:,.2f}").font = success_font
            current_row += 1
            
            # 無效轉化統計 - 總是顯示，即使為0，使用紅色字體
            invalid_mask = df['Status'].str.lower().isin(['invalid', 'rejected'])
            invalid_count = invalid_mask.sum()
            invalid_amount = df[invalid_mask]['USD Sale Amount'].sum() if invalid_count > 0 else 0.0
            
            # 使用紅色字體顯示無效轉化統計
            invalid_font = Font(size=10, color="FF0000")  # 紅色
            ws.cell(row=current_row, column=1, value=f"⚠️ Total Conversions (Invalid/Rejected): {invalid_count:,}").font = invalid_font
            current_row += 1
            ws.cell(row=current_row, column=1, value=f"⚠️ Total Sale Amount (USD) (Invalid/Rejected): ${invalid_amount:,.2f}").font = invalid_font
            current_row += 1
        
        current_row += 1  # 空一行
        return current_row
    
    def _write_data_to_sheet(self, ws, df: pd.DataFrame, start_row: int,
                           header_font, header_fill, header_alignment, 
                           data_alignment, number_alignment):
        """將數據寫入Excel sheet"""
        # 寫入標題行
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 寫入數據行
        data_start_row = start_row + 1
        for row_idx, (_, row) in enumerate(df.iterrows(), data_start_row):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 設置格式
                if isinstance(value, (int, float)):
                    cell.alignment = number_alignment
                    if 'USD' in headers[col_idx-1]:
                        cell.number_format = '"$"#,##0.00'
                else:
                    cell.alignment = data_alignment
    
    def _auto_adjust_columns(self, ws):
        """自動調整列寬"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    async def _send_all_partner_emails(self, excel_files: List[str], self_email: bool = False, 
                                     start_date: Optional[str] = None, end_date: Optional[str] = None) -> Dict[str, Any]:
        """發送所有Partner的郵件"""
        try:
            print_step("Partner邮件发送", "开始按Partner分别发送转换报告邮件")
            
            # 初始化郵件發送器時添加調試信息
            try:
                logger.info("🔧 正在初始化郵件發送器...")
                if not hasattr(self, 'email_sender') or self.email_sender is None:
                    self.email_sender = EmailSender()
                    logger.info("✅ 郵件發送器初始化成功")
                
                # 測試SMTP連接
                logger.info("🔍 測試SMTP連接...")
                if not self.email_sender._test_smtp_connection():
                    logger.error("❌ SMTP連接測試失敗")
                    return {
                        'success': False,
                        'error': 'SMTP連接失敗',
                        'total_sent': 0,
                        'total_failed': len(excel_files),
                        'partner_results': {}
                    }
                else:
                    logger.info("✅ SMTP連接測試成功")
                    
            except Exception as init_error:
                logger.error(f"❌ 郵件發送器初始化失敗: {init_error}")
                return {
                    'success': False,
                    'error': f'郵件發送器初始化失敗: {init_error}',
                    'total_sent': 0,
                    'total_failed': len(excel_files),
                    'partner_results': {}
                }
            
            partner_results = {}
            total_sent = 0
            total_failed = 0
            
            for excel_file in excel_files:
                try:
                    # 從文件名提取Partner名稱
                    partner_name = os.path.basename(excel_file).split('_')[0]
                    print_step("邮件发送", f"📧 正在发送 {partner_name} 的邮件...")
                    
                    # 發送郵件
                    # 🎯 修復：優先使用傳入的日期參數，確保郵件日期與實際數據日期範圍一致
                    if start_date and end_date:
                        # 使用傳入的日期參數
                        report_date = end_date
                        start_date_param = start_date
                        logger.info(f"📅 使用傳入的日期範圍: {start_date} to {end_date}")
                    else:
                        # 備用方案：從文件名中提取日期
                        extracted_date = self._extract_date_from_filename(excel_file)
                        if extracted_date:
                            report_date = extracted_date
                            start_date_param = extracted_date
                            logger.info(f"📅 使用文件名中的日期: {extracted_date}")
                        else:
                            # 最後備用方案：使用當前日期
                            report_date = datetime.now().strftime('%Y-%m-%d')
                            start_date_param = report_date
                            logger.info(f"📅 使用當前日期: {report_date}")
                    
                    result = self.email_sender.send_partner_reports(
                        partner_summary={partner_name: {'file_path': excel_file}},
                        report_date=report_date,
                        start_date=start_date_param,
                        self_email=self_email
                    )
                    
                    if result and result.get('success'):
                        partner_results[partner_name] = {'success': True}
                        total_sent += 1
                        print_step("邮件发送", f"✅ {partner_name} 邮件发送成功")
                    else:
                        error_msg = result.get('error', '发送失败') if result else '无响应'
                        partner_results[partner_name] = {'success': False, 'error': error_msg}
                        total_failed += 1
                        print_step("邮件发送", f"❌ {partner_name} 邮件发送失败: {error_msg}")
                        
                except Exception as e:
                    # 如果出錯，確保partner_name有值
                    if 'partner_name' not in locals():
                        partner_name = f"Unknown_{os.path.basename(excel_file)}"
                    partner_results[partner_name] = {'success': False, 'error': str(e)}
                    total_failed += 1
                    print_step("邮件发送", f"❌ {partner_name} 邮件发送异常: {e}")
            
            # 匯總結果
            overall_success = total_failed == 0
            status_msg = f"⚠️ 邮件发送完成: 成功 {total_sent} 封，失败 {total_failed} 封"
            if overall_success:
                status_msg = f"✅ 邮件发送完成: 全部 {total_sent} 封邮件发送成功"
            
            print_step("邮件发送", status_msg)
            
            return {
                'success': overall_success,
                'total_sent': total_sent,
                'total_failed': total_failed,
                'partner_results': partner_results
            }
            
        except Exception as e:
            error_msg = f"邮件发送流程异常: {e}"
            logger.error(f"❌ {error_msg}")
            return {
                'success': False,
                'error': error_msg,
                'total_sent': 0,
                'total_failed': len(excel_files),
                'partner_results': {}
            }
    
    async def _send_single_partner_email(self, partner_summary: PartnerSummary, self_email: bool):
        """發送單個Partner的郵件"""
        try:
            logger.info(f"📧 準備發送 {partner_summary.partner_name} 的郵件，file_path: {partner_summary.file_path}")
            if not partner_summary.file_path:
                logger.warning(f"⚠️ {partner_summary.partner_name} 的file_path為None，跳過郵件發送")
                return
                
            await self._send_all_partner_emails([partner_summary.file_path], self_email)
            
        except Exception as e:
            logger.error(f"❌ 發送 {partner_summary.partner_name} 郵件失敗: {e}")
    
    async def _upload_all_partner_feishu(self, excel_files: List[str]):
        """上傳所有Partner文件到飛書"""
        try:
            logger.info(f"📱 開始上傳 {len(excel_files)} 個Partner文件到飛書...")
            
            upload_result = self.feishu_uploader.upload_files(excel_files)
            
            if upload_result.get('success'):
                logger.info(f"✅ 飛書上傳成功: {upload_result['success_count']} 個文件")
                for file_path in excel_files:
                    filename = os.path.basename(file_path)
                    logger.info(f"   📄 已上傳: {filename}")
            else:
                logger.warning(f"⚠️ 飛書上傳部分失敗: 成功 {upload_result.get('success_count', 0)} 個，失敗 {upload_result.get('failed_count', 0)} 個")
                
        except Exception as e:
            logger.error(f"❌ 飛書上傳失敗: {e}")
    
    async def _upload_single_partner_feishu(self, excel_files: List[str], partner_name: str):
        """上傳單個Partner文件到飛書"""
        try:
            await self._upload_all_partner_feishu(excel_files)
        except Exception as e:
            logger.error(f"❌ 上傳 {partner_name} 文件到飛書失敗: {e}") 
    
    async def _print_final_summary(self, df: pd.DataFrame, partner_summaries: List[PartnerSummary], import_file_path: str):
        """生成並打印最終的重點總結報告"""
        try:
            print_step("", "")
            print_step("🎯 Report Summary", "SUMMARY")
            print_step("=" * 80, "")
            
            # 1. 獲取基本信息
            partners_list = [s.partner_name for s in partner_summaries if s.total_records > 0]
            date_range = self._extract_date_range_from_filename(import_file_path)
            
            # 2. 獲取原始數據和Mockup信息
            original_stats = await self._get_original_data_stats()
            
            # 獲取Partner特定的mockup倍數
            from config import get_partner_mockup_multiplier
            partner_mockup_info = {}
            
            # 為每個Partner獲取mockup配置
            for partner_name in partners_list:
                if partner_name.upper() in ['RAMPUP', 'DEEPLEAPER', 'TESTPARTNER', 'MKK', 'MP', 'FTK', 'BYTEC']:
                    mockup_multiplier = get_partner_mockup_multiplier(partner_name.upper())
                    partner_mockup_info[partner_name] = mockup_multiplier
                else:
                    # 使用默認倍數
                    mockup_multiplier = getattr(config, 'MOCKUP_MULTIPLIER', 0.9)
                    partner_mockup_info[partner_name] = mockup_multiplier
            
            # 如果只有一個Partner，使用其mockup倍數；否則使用默認倍數
            if len(partners_list) == 1:
                mockup_multiplier = partner_mockup_info[partners_list[0]]
            else:
                mockup_multiplier = getattr(config, 'MOCKUP_MULTIPLIER', 0.9)
            
            # 3. 計算當前數據統計
            total_conversions = len(df)
            pending_approved_df = df[df['Status'].isin(['Pending', 'Approved'])]
            invalid_rejected_df = df[df['Status'].isin(['Invalid', 'Rejected'])]
            
            total_pending_approved = len(pending_approved_df)
            total_invalid_rejected = len(invalid_rejected_df)
            total_amount_pending_approved = pending_approved_df['USD Sale Amount'].sum()
            total_amount_invalid_rejected = invalid_rejected_df['USD Sale Amount'].sum()
            
            # 4. 輸出格式化報告
            print_step(f"Partner: {', '.join(partners_list)}", "")
            print_step(f"Date Range: {date_range}", "")
            
            if original_stats:
                print_step(f"Original Total Conversions (All Status): {original_stats['total_conversions']:,}", "")
                print_step(f"Original Sale Amount (USD) (All Status): ${original_stats['total_amount']:,.2f}", "")
            
            print_step(f"Mockup: {mockup_multiplier} (from @config.py)", "")
            
            # 添加Partner mockup配置詳細信息
            print_step("", "")
            print_step("🎯 Partner Mockup 配置詳情:", "")
            for partner_name, multiplier in partner_mockup_info.items():
                print_step(f"   📊 {partner_name}: {multiplier} ({multiplier * 100}%)", "")
            
            print_step("", "")
            
            print_step(f"Total Conversions (All Status): {total_conversions:,}", "")
            print_step(f"✅ Total Conversions (Pending/Approved): {total_pending_approved:,}", "")
            print_step(f"✅ Total Sale Amount (USD) (Pending/Approved): ${total_amount_pending_approved:,.2f}", "")
            
            if total_invalid_rejected > 0:
                print_step(f"⚠️ Total Conversions (Invalid/Rejected): {total_invalid_rejected:,}", "")
                print_step(f"⚠️ Total Sale Amount (USD) (Invalid/Rejected): ${total_amount_invalid_rejected:,.2f}", "")
            else:
                print_step(f"⚠️ Total Conversions (Invalid/Rejected): 0", "")
                print_step(f"⚠️ Total Sale Amount (USD) (Invalid/Rejected): $0.00", "")
            
            print_step("=" * 80, "")
            
            # 添加Partner mockup前後情況總結
            await self._print_partner_mockup_summary(partner_mockup_info, original_stats, total_amount_pending_approved)
            
        except Exception as e:
            logger.error(f"❌ 生成重點總結失敗: {e}")
    
    async def _print_partner_mockup_summary(self, partner_mockup_info: Dict[str, float], original_stats: Optional[Dict[str, Any]], current_amount: float):
        """打印Partner mockup前後情況總結"""
        try:
            print_step("", "")
            print_step("📊 Partner Mockup 前後情況總結", "SUMMARY")
            print_step("=" * 80, "")
            
            # 計算原始金額（如果可用）
            if original_stats and 'total_amount' in original_stats:
                original_amount = original_stats['total_amount']
                print_step(f"💰 原始總金額 (USD): ${original_amount:,.2f}", "")
                print_step(f"💰 當前總金額 (USD): ${current_amount:,.2f}", "")
                
                # 計算變化
                amount_change = current_amount - original_amount
                change_percentage = (amount_change / original_amount * 100) if original_amount > 0 else 0
                print_step(f"📈 金額變化: ${amount_change:+,.2f} ({change_percentage:+.2f}%)", "")
            else:
                print_step(f"💰 當前總金額 (USD): ${current_amount:,.2f}", "")
            
            print_step("", "")
            print_step("🎯 Partner Mockup 配置影響:", "")
            
            # 分析每個Partner的mockup影響
            total_original_estimate = 0
            total_adjusted_estimate = 0
            
            for partner_name, multiplier in partner_mockup_info.items():
                if multiplier != 1.0:  # 只顯示有調整的Partner
                    adjustment = (multiplier - 1) * 100
                    print_step(f"   📊 {partner_name}: {multiplier} ({adjustment:+.1f}% 調整)", "")
                    
                    # 估算該Partner的原始金額（基於當前金額和倍數）
                    if current_amount > 0:
                        # 假設所有Partner的數據比例相同，估算該Partner的原始金額
                        partner_original_estimate = current_amount / multiplier
                        partner_adjusted_estimate = current_amount
                        total_original_estimate += partner_original_estimate
                        total_adjusted_estimate += partner_adjusted_estimate
                else:
                    print_step(f"   📊 {partner_name}: {multiplier} (無調整)", "")
            
            print_step("", "")
            print_step("📈 Mockup 影響估算:", "")
            if total_original_estimate > 0:
                print_step(f"   💰 估算原始總金額: ${total_original_estimate:,.2f}", "")
                print_step(f"   💰 調整後總金額: ${total_adjusted_estimate:,.2f}", "")
                print_step(f"   📊 總調整幅度: {((total_adjusted_estimate / total_original_estimate - 1) * 100):+.2f}%", "")
            
            print_step("", "")
            print_step("🔧 Mockup 配置來源: @config.py", "")
            print_step("   - 所有Partner的mockup倍數都從config.py的PARTNER_SOURCES_MAPPING獲取", "")
            print_step("   - 使用get_partner_mockup_multiplier()函數獲取特定Partner的倍數", "")
            
            # 打印所有Partner的配置
            print_step("", "")
            print_step("📋 所有Partner Mockup 配置:", "")
            for partner_name, multiplier in partner_mockup_info.items():
                status = "✅ 已調整" if multiplier != 1.0 else "➖ 無調整"
                print_step(f"   📊 {partner_name}: {multiplier} ({multiplier * 100}%) - {status}", "")
            
            print_step("", "")
            print_step("🎯 重點提醒:", "")
            print_step("   - FTK Partner: 100% (無調整)", "")
            print_step("   - 其他Partner: 80% (20% 調整)", "")
            print_step("   - ByteC Partner: 80% (20% 調整)", "")
            
        except Exception as e:
            logger.error(f"❌ 打印Partner mockup總結失敗: {e}")
            print_step(f"❌ 打印Partner mockup總結失敗: {e}", "")
    
    def _get_date_range_for_filename(self, start_date: Optional[str], end_date: Optional[str], import_file_path: str) -> str:
        """獲取檔名中的日期範圍"""
        try:
            # 優先使用傳入的 start_date 和 end_date
            if start_date and end_date:
                return f"{start_date}_to_{end_date}"
            
            # 其次從文件名中提取
            filename_date_range = self._extract_date_range_from_filename(import_file_path)
            if filename_date_range != "unknown_date_range":
                return filename_date_range
                
            # 最後使用當前日期作為備用
            today_str = datetime.now().strftime("%Y-%m-%d")
            return f"{today_str}_to_{today_str}"
        except Exception as e:
            logger.warning(f"⚠️ 獲取日期範圍失敗: {e}")
            return "unknown_date_range"
    
    def _extract_date_from_filename(self, file_path: str) -> Optional[str]:
        """從文件名中提取單個日期（用於郵件日期）"""
        try:
            filename = os.path.basename(file_path)
            
            # 匹配8位数字日期格式（YYYYMMDD）
            match = re.search(r'(\d{8})', filename)
            if match:
                date_str = match.group(1)
                # 格式化為 YYYY-MM-DD
                formatted_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
                logger.info(f"📅 從文件名提取日期: {formatted_date}")
                return formatted_date
            
            # 備用方案：使用昨天日期
            yesterday = datetime.now() - timedelta(days=1)
            date_str = yesterday.strftime("%Y-%m-%d")
            logger.info(f"📅 使用備用日期: {date_str}")
            return date_str
            
        except Exception as e:
            logger.warning(f"⚠️ 提取日期失敗: {e}")
            return None
    
    def _extract_date_range_from_filename(self, import_file_path: str) -> str:
        """從文件路徑中提取日期範圍"""
        try:
            # 嘗試從文件名中提取日期，支持多種格式：
            # 1. 'publisher-conversion-report--LXTlT9i5-20250725.csv'
            # 2. 'DMP_temp_DeepLeaper_20250804_215355.xlsx'
            filename = os.path.basename(import_file_path)
            
            # 匹配8位数字日期格式（YYYYMMDD）
            match = re.search(r'(\d{8})', filename)
            if match:
                date_str = match.group(1)
                # 格式化為 YYYY-MM-DD
                formatted_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
                return f"{formatted_date}_to_{formatted_date}"
            
            # 備用方案：使用當前日期
            yesterday = datetime.now() - timedelta(days=1)
            date_str = yesterday.strftime("%Y-%m-%d")
            return f"{date_str}_to_{date_str}"
            
        except Exception as e:
            logger.warning(f"⚠️ 提取日期範圍失敗: {e}")
            return "unknown_date_range"
    
    async def _get_original_data_stats(self) -> Optional[Dict[str, Any]]:
        """嘗試從DMP Agent日誌或其他來源獲取原始數據統計"""
        try:
            # 方案1: 從最新的Data Input Agent輸出文件獲取原始數據
            import glob
            
            # 查找最新的Passthrough文件
            passthrough_files = glob.glob('output/Passthrough_*.xlsx')
            if passthrough_files:
                latest_passthrough = max(passthrough_files, key=os.path.getctime)
                df_original = pd.read_excel(latest_passthrough)
                
                # 獲取原始統計（mockup前）
                original_total_conversions = len(df_original)
                # 注意: Data Input Agent輸出的是 'Sale Amount (USD)'，還未被mockup調整
                original_total_amount = df_original['Sale Amount (USD)'].sum()
                
                return {
                    'total_conversions': original_total_conversions,
                    'total_amount': original_total_amount
                }
            
            # 方案2: 如果找不到原始文件，返回None
            return None
            
        except Exception as e:
            logger.warning(f"⚠️ 獲取原始數據統計失敗: {e}")
            return None