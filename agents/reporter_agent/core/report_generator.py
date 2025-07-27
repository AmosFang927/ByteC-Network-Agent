#!/usr/bin/env python3
"""
报表生成器
从数据库生成Excel报表，并发送到飞书和邮件
增强版：包含詳細的性能監控和時間統計
"""

import os
import sys
import asyncio
import pandas as pd
import time
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
import logging
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 强制日志输出到前台
os.environ['PYTHONUNBUFFERED'] = '1'

# 配置日志强制输出到前台
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    stream=sys.stdout,
    force=True
)

# 确保所有logger都输出到stdout
for handler in logging.root.handlers:
    handler.setStream(sys.stdout)

# 强制刷新stdout
sys.stdout.flush()

# 导入ByteC-Network-Agent的现有模块
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
import config
from modules.feishu_uploader import FeishuUploader
from agents.data_output_agent.email_sender import EmailSender
from utils.logger import print_step
from .database import PostbackDatabase, PartnerSummary
from shared.utils.summary_formatter import generate_unified_summary, format_summary_for_display

logger = logging.getLogger(__name__)

class ReportGenerator:
    """报表生成器 - 增強版本包含性能監控"""
    
    def __init__(self, output_dir: str = "output", 
                 global_email_disabled: bool = False):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # 性能統計追蹤
        self.performance_stats = {
            'total_start_time': None,
            'steps': [],
            'database_query_time': 0,
            'excel_generation_time': 0,
            'email_send_time': 0,
            'feishu_upload_time': 0,
            'total_records_processed': 0
        }
        
        # 初始化数据库连接
        self.db = PostbackDatabase()
        
        # 初始化飞书上传和邮件发送器
        self.feishu_uploader = FeishuUploader()
        self.email_sender = EmailSender(global_email_disabled=global_email_disabled)
        
        # 邮件配置
        self.partner_email_mapping = {
            'InvolveAsia': ['partners@involveasia.com'],
            'Rector': ['rector@partners.com'],
            'DeepLeaper': ['deepleaper@partners.com'],
            'ByteC': ['bytec@partners.com'],
            'RAMPUP': ['rampup@partners.com'],
            'ALL': ['AmosFang927@gmail.com']
        }
        
        self.partner_email_enabled = {
            'InvolveAsia': True,
            'Rector': True,
            'DeepLeaper': True,
            'ByteC': True,
            'RAMPUP': True,
            'ALL': True
        }
    
    def _start_step_timer(self, step_name: str) -> float:
        """開始計時一個步驟"""
        start_time = time.time()
        logger.info(f"⏱️  開始執行: {step_name}")
        return start_time
    
    def _end_step_timer(self, step_name: str, start_time: float) -> float:
        """結束計時一個步驟並記錄"""
        duration = time.time() - start_time
        self.performance_stats['steps'].append({
            'name': step_name,
            'duration': duration,
            'timestamp': datetime.now()
        })
        logger.info(f"✅ 完成執行: {step_name} (耗時 {duration:.2f}秒)")
        return duration
    
    def _log_performance_summary(self):
        """記錄性能總結"""
        if not self.performance_stats['total_start_time']:
            return
            
        total_time = time.time() - self.performance_stats['total_start_time']
        
        logger.info("=" * 60)
        logger.info("📊 **性能統計總結**")
        logger.info("=" * 60)
        logger.info(f"🕐 總執行時間: {total_time:.2f}秒")
        logger.info(f"📊 處理記錄數: {self.performance_stats['total_records_processed']:,}")
        
        if self.performance_stats['total_records_processed'] > 0:
            records_per_sec = self.performance_stats['total_records_processed'] / total_time
            logger.info(f"⚡ 處理速度: {records_per_sec:.2f} 記錄/秒")
        
        logger.info("📋 各步驟耗時明細:")
        for step in self.performance_stats['steps']:
            percentage = (step['duration'] / total_time) * 100
            logger.info(f"   • {step['name']}: {step['duration']:.2f}秒 ({percentage:.1f}%)")
        
        logger.info("=" * 60)
    
    async def generate_partner_report(self, partner_name: str = "ALL",
                                    start_date: datetime = None,
                                    end_date: datetime = None,
                                    send_email: bool = True,
                                    upload_feishu: bool = True,
                                    self_email: bool = False,
                                    limit: Optional[int] = None) -> Dict[str, Any]:
        """
        生成Partner报表
        
        Args:
            partner_name: Partner名称 (ALL, InvolveAsia, Rector, DeepLeaper, ByteC, RAMPUP)
            start_date: 开始日期
            end_date: 结束日期
            send_email: 是否发送邮件
            upload_feishu: 是否上传到飞书
            limit: 限制拉取的记录数量
            
        Returns:
            Dict[str, Any]: 报表生成结果
        """
        # 開始總體計時
        self.performance_stats['total_start_time'] = time.time()
        overall_start = self._start_step_timer("完整報表生成流程")
        
        try:
            logger.info(f"🚀 开始生成 {partner_name} 报表")
            
            # 设置默认日期范围（过去7天）
            if not end_date:
                end_date = datetime.now()
            if not start_date:
                start_date = end_date - timedelta(days=7)
            
            # 记录limit设置
            limit_info = f" (限制: {limit} 条)" if limit else ""
            logger.info(f"🔍 查询数据: {start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}{limit_info}")
            
            # 步驟 1: 數據庫查詢
            db_start = self._start_step_timer("數據庫查詢")
            df = await self.db.get_conversion_dataframe(partner_name, start_date, end_date, limit=limit)
            self.performance_stats['database_query_time'] = self._end_step_timer("數據庫查詢", db_start)
            
            if df.empty:
                logger.warning(f"⚠️ 没有找到 {partner_name} 的转化数据")
                print(f"📋 提示: {partner_name} 在 {start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')} 期间没有转化数据，將生成空報表並繼續執行")
                
                # 创建空的partner_summaries
                partner_summaries = []
                self.performance_stats['total_records_processed'] = 0
            else:
                # 如果设置了limit且达到限制，显示信息
                if limit and len(df) >= limit:
                    logger.info(f"📊 已达到数据拉取限制: {len(df)} 条记录 (设置限制: {limit})")
                    print(f"⏹️ 数据收取已停止: 达到设置的 {limit} 条记录限制")
                
                self.performance_stats['total_records_processed'] = len(df)
                
                # 步驟 2: Partner 彙總查詢
                summary_start = self._start_step_timer("Partner 彙總查詢")
                partner_summaries = await self.db.get_partner_summary(partner_name, start_date, end_date, limit=limit)
                self._end_step_timer("Partner 彙總查詢", summary_start)
            
            # 步驟 3: 生成Excel文件
            excel_start = self._start_step_timer("Excel文件生成")
            excel_files = await self._generate_excel_files(df, partner_summaries, partner_name, start_date, end_date)
            self.performance_stats['excel_generation_time'] = self._end_step_timer("Excel文件生成", excel_start)
            
            # 計算總金額（處理空數據情況，排除 invalid/rejected）
            if not partner_summaries:
                total_amount = 0.0
                total_records = 0
            else:
                # 直接從 partner_summaries 計算總和，確保一致性
                total_amount = sum(summary.total_amount for summary in partner_summaries)
                total_records = sum(summary.total_records for summary in partner_summaries)
                logger.info(f"🔍 [MONITOR] generate_partner_report 統一總金額: ${total_amount:,.2f} (總記錄數: {total_records})")
            
            # 獲取無效轉化統計
            invalid_stats = df.attrs.get('invalid_stats', {'invalid_count': 0, 'invalid_amount': 0.0}) if not df.empty else {'invalid_count': 0, 'invalid_amount': 0.0}
            
            result = {
                'success': True,
                'partner_name': partner_name,
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d'),
                'total_records': total_records,  # 使用有效記錄數
                'total_amount': total_amount,
                'invalid_stats': invalid_stats,  # 添加無效轉化統計
                'excel_files': excel_files,
                'partner_summaries': [summary.to_dict() for summary in partner_summaries] if partner_summaries else []
            }
            
            # 步驟 4: 飞书上传
            if upload_feishu and excel_files:
                feishu_start = self._start_step_timer("飛書文件上傳")
                feishu_result = await self._upload_to_feishu(excel_files)
                self.performance_stats['feishu_upload_time'] = self._end_step_timer("飛書文件上傳", feishu_start)
                result['feishu_upload'] = feishu_result
            
            # 步驟 5: 邮件发送
            if send_email and excel_files:
                email_start = self._start_step_timer("郵件發送")
                email_result = await self._send_emails(partner_summaries, excel_files, start_date, end_date, self_email)
                self.performance_stats['email_send_time'] = self._end_step_timer("郵件發送", email_start)
                result['email_result'] = email_result
            
            # 完成總體計時並生成性能總結
            self._end_step_timer("完整報表生成流程", overall_start)
            self._log_performance_summary()
            
            logger.info(f"✅ {partner_name} 报表生成完成")
            return result
            
        except Exception as e:
            # 即使出錯也記錄性能統計
            if self.performance_stats['total_start_time']:
                self._end_step_timer("完整報表生成流程", overall_start)
                self._log_performance_summary()
            
            logger.error(f"❌ 生成报表失败: {e}")
            return {
                'success': False,
                'error': str(e),
                'partner_name': partner_name
            }
    
    async def _generate_excel_files(self, df: pd.DataFrame, 
                                  partner_summaries: List[PartnerSummary],
                                  partner_name: str, 
                                  start_date: datetime,
                                  end_date: datetime) -> List[str]:
        """生成Excel文件"""
        excel_files = []
        
        try:
            if partner_name.upper() == 'ALL':
                # 为每个Partner生成单独的文件（即使沒有數據也要生成）
                for summary in partner_summaries:
                    partner_df = df[df['Partner'] == summary.partner_name]
                    
                    # 即使沒有數據也生成Excel文件
                    file_path = await self._create_excel_file(
                        partner_df, 
                        summary.partner_name, 
                        start_date, 
                        end_date,
                        [summary]  # 傳遞對應的 partner_summary
                    )
                    excel_files.append(file_path)
                    
                    # 为每个Partner的汇总添加文件路径
                    summary.file_path = file_path
                    
                    # 記錄生成信息
                    record_count = len(partner_df)
                    logger.info(f"📄 生成 {summary.partner_name} Excel文件: {record_count} 條記錄")
                
                # 根據配置決定是否生成总汇总文件
                if getattr(config, 'GENERATE_ALLPARTNERS_REPORT', False):
                    main_file = await self._create_excel_file(df, "AllPartners", start_date, end_date, partner_summaries)
                    excel_files.insert(0, main_file)  # 插入到第一个位置
                    logger.info("📄 已生成 AllPartners 總汇总文件")
                else:
                    logger.info("📄 跳過 AllPartners 總汇总文件生成（配置已禁用）")
                
            else:
                # 生成单个Partner的文件（即使是空數據也要生成）
                # 找到对应的 partner_summary
                target_summary = None
                for summary in partner_summaries:
                    if summary.partner_name == partner_name:
                        target_summary = summary
                        break
                
                file_path = await self._create_excel_file(
                    df, 
                    partner_name, 
                    start_date, 
                    end_date, 
                    [target_summary] if target_summary else None
                )
                excel_files.append(file_path)
                
                # 为汇总添加文件路径
                if partner_summaries:
                    partner_summaries[0].file_path = file_path
            
            logger.info(f"✅ 成功生成 {len(excel_files)} 个Excel文件")
            return excel_files
            
        except Exception as e:
            logger.error(f"❌ 生成Excel文件失败: {e}")
            raise
    
    def _add_summary_header(self, ws, partner_name: str, start_date: datetime, 
                           end_date: datetime, total_conversions: int, 
                           total_amount: float, current_row: int = 1, 
                           invalid_conversions: int = 0, invalid_amount: float = 0.0,
                           df: pd.DataFrame = None) -> int:
        """添加匯總信息區塊到Excel工作表 - 使用统一Summary格式化器"""
        # 樣式設置
        title_font = Font(bold=True, size=14, color="FFFFFF")
        title_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
        info_font = Font(size=11)
        warning_font = Font(bold=True, color="FF0000")  # 紅色警告字體
        success_font = Font(bold=True, color="008000")  # 綠色成功字體
        
        # 標題行 - 使用統一格式
        ws.cell(row=current_row, column=1, value=f"📊 {partner_name} 报告摘要").font = title_font
        ws.cell(row=current_row, column=1).fill = title_fill
        current_row += 1
        
        # 使用统一Summary格式化器生成Summary
        try:
            # 生成统一Summary
            unified_summary = generate_unified_summary(
                partner_name=partner_name,
                start_date=start_date.strftime('%Y-%m-%d'),
                end_date=end_date.strftime('%Y-%m-%d'),
                df=df,
                total_records=total_conversions,
                total_amount=total_amount,
                sources=[]  # 在Excel中暂时不显示sources
            )
            
            # 日期範圍
            ws.cell(row=current_row, column=1, value=f"Date Range: {unified_summary['date_range']}").font = info_font
            current_row += 1
            
            # 總轉化數（所有狀態）
            ws.cell(row=current_row, column=1, value=f"Total Conversions (All Status): {unified_summary['total_all_conversions']:,}").font = info_font
            current_row += 1
            
            # Total Conversions (Pending/Approved)
            ws.cell(row=current_row, column=1, value=f"✅ Total Conversions (Pending/Approved): {unified_summary['pending_approved_count']:,}").font = success_font
            current_row += 1
            
            # Total Sale Amount (USD) (Pending/Approved)
            ws.cell(row=current_row, column=1, value=f"✅ Total Sale Amount (USD) (Pending/Approved): {unified_summary['pending_approved_amount']}").font = success_font
            current_row += 1
            
            # 總是顯示 Invalid/Rejected 記錄（即使值為0）
            ws.cell(row=current_row, column=1, value=f"⚠️ Total Conversions (Invalid/Rejected): {unified_summary['invalid_rejected_count']:,}").font = warning_font
            current_row += 1
            ws.cell(row=current_row, column=1, value=f"⚠️ Total Sale Amount (USD) (Invalid/Rejected): {unified_summary['invalid_rejected_amount']}").font = warning_font
            current_row += 1
            
            # 添加Sources信息
            sources_list = unified_summary.get('sources_list', 'N/A')
            ws.cell(row=current_row, column=1, value=f"Sources: {sources_list}").font = info_font
            current_row += 1
            
            logger.info(f"✅ Excel Summary使用统一格式: {partner_name}, 有效转化: {unified_summary['pending_approved_count']}, 金额: {unified_summary['pending_approved_amount']}")
            
        except Exception as e:
            logger.error(f"❌ 生成统一Summary失败: {e}")
            # 使用原有逻辑作为备用
            date_range = f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
            ws.cell(row=current_row, column=1, value=f"Date Range: {date_range}").font = info_font
            current_row += 1
            
            # 總轉化數
            ws.cell(row=current_row, column=1, value=f"Total Conversions: {total_conversions:,}").font = info_font
            current_row += 1
            
            # 總銷售金額
            ws.cell(row=current_row, column=1, value=f"Total Sale Amount (USD): ${total_amount:,.2f}").font = info_font
            current_row += 1
        
        # 報告創建時間
        from datetime import datetime
        report_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.cell(row=current_row, column=1, value=f"Report Created Time: {report_time}").font = info_font
        current_row += 2  # 空一行
        
        return current_row

    async def _create_excel_file(self, df: pd.DataFrame, 
                               partner_name: str,
                               start_date: datetime,
                               end_date: datetime,
                               partner_summaries: List[PartnerSummary] = None) -> str:
        """創建Excel文件"""
        try:
            # 創建文件名
            clean_partner_name = self._clean_sheet_name(partner_name)
            filename = f"{clean_partner_name}_ConversionReport_{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}.xlsx"
            file_path = os.path.join(self.output_dir, filename)
            
            # 創建工作簿
            wb = Workbook()
            
            # 刪除默認工作表
            wb.remove(wb.active)
            
            # 樣式設置
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            data_alignment = Alignment(horizontal="left", vertical="center")
            number_alignment = Alignment(horizontal="right", vertical="center")
            
            # 創建總表工作表
            clean_partner_name = self._clean_sheet_name(partner_name)
            summary_ws = wb.create_sheet(clean_partner_name, 0)
            
            if not df.empty:
                # 初始化为0，以防没有找到匹配的summary
                excel_total_amount = 0.0
                excel_total_conversions = 0

                if partner_name.upper() == 'ALL':
                    # For 'ALL' partner, sum up all summaries
                    if partner_summaries:
                        excel_total_amount = sum(float(s.total_amount) for s in partner_summaries)
                        excel_total_conversions = sum(s.total_records for s in partner_summaries)
                        logger.info(f"🔍 [MONITOR] Excel总表 (ALL) 使用partner_summaries总金额: ${excel_total_amount:,.2f} (记录数: {excel_total_conversions})")
                        logger.debug(f"🔬 [DEBUG] All Partner Summaries (Excel): {[s.to_dict() for s in partner_summaries]}")
                        # 添加日誌
                        logger.debug(f"[DEBUG] After calculation for ALL: Total Amount = {excel_total_amount}, Total Conversions = {excel_total_conversions}")
                else:
                    # For specific partner, find the matching summary
                    target_summary = next((s for s in partner_summaries if s.partner_name == partner_name), None)
                    if target_summary:
                        excel_total_amount = float(target_summary.total_amount)
                        excel_total_conversions = target_summary.total_records
                        logger.info(f"🔍 [MONITOR] Excel总表 使用partner_summaries金额: ${excel_total_amount:,.2f} (记录数: {excel_total_conversions})")
                        logger.debug(f"🔬 [DEBUG] Partner Summary详情 (Excel): {target_summary}")
                        # 添加日誌
                        logger.debug(f"[DEBUG] After calculation for {partner_name}: Total Amount = {excel_total_amount}, Total Conversions = {excel_total_conversions}")
                    else:
                        logger.warning(f"⚠️ [WARNING] 未在 partner_summaries 中找到 {partner_name} 的數據。Excel 金額將顯示為 0。")

                # 獲取無效轉化統計 (保持原樣，因為它獨立於總金額計算)
                invalid_stats = df.attrs.get('invalid_stats', {'invalid_count': 0, 'invalid_amount': 0.0})
                invalid_count = invalid_stats.get('invalid_count', 0)
                invalid_amount = invalid_stats.get('invalid_amount', 0.0)

                # 添加匯總信息區塊 - 使用我們計算的 excel_total_conversions 和 excel_total_amount
                current_row = self._add_summary_header(
                    summary_ws, partner_name, start_date, end_date, 
                    excel_total_conversions, excel_total_amount, 1, invalid_count, invalid_amount, df
                )
                
                # 寫入總表數據（所有數據）
                headers = list(df.columns)
                
                # 寫入標題行
                for col_idx, header in enumerate(headers, 1):
                    cell = summary_ws.cell(row=current_row, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 寫入數據行
                data_start_row = current_row + 1
                for row_idx, (_, row) in enumerate(df.iterrows(), data_start_row):
                    for col_idx, value in enumerate(row, 1):
                        # 清理Excel不支持的字符
                        from utils.excel_character_cleaner import clean_for_excel
                        cleaned_value = clean_for_excel(value)
                        cell = summary_ws.cell(row=row_idx, column=col_idx, value=cleaned_value)
                        
                        # 設置數據格式
                        if isinstance(value, (int, float)):
                            cell.alignment = number_alignment
                            if 'USD Sale Amount' in headers[col_idx-1]:
                                cell.number_format = '"$"#,##0.00'
                        else:
                            cell.alignment = data_alignment
                
                # 自動調整列寬
                for column in summary_ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    summary_ws.column_dimensions[column_letter].width = adjusted_width
                
                # 2. 然後按 Source 分組創建各個 Sheet
                if not df.empty and 'Source' in df.columns:
                    sources = df['Source'].unique()
                    
                    for source in sources:
                        if pd.isna(source) or source == '':
                            sheet_name = "Unknown"
                        else:
                            # 直接使用原始Source值，只清理不符合Excel要求的字符
                            sheet_name = str(source).replace('/', '_').replace('\\', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_').replace('?', '_')
                            # 限制長度
                            if len(sheet_name) > 31:
                                sheet_name = sheet_name[:31]
                        
                        # 檢查是否已存在同名工作表（改進的重名處理邏輯）
                        original_sheet_name = sheet_name
                        counter = 1
                        while sheet_name in wb.sheetnames:
                            # 如果與Partner主sheet重名，使用Source前綴區分
                            if sheet_name == clean_partner_name:
                                sheet_name = f"Source_{original_sheet_name}"
                            else:
                                # 其他重名情況，使用數字后綴
                                sheet_name = f"{original_sheet_name}_{counter}"
                                counter += 1
                            
                            # 確保長度不超過31字符
                            if len(sheet_name) > 31:
                                if sheet_name.startswith("Source_"):
                                    sheet_name = f"Source_{original_sheet_name[:24]}"
                                else:
                                    sheet_name = f"{original_sheet_name[:28]}_{counter-1}"
                        
                        ws = wb.create_sheet(sheet_name)
                        
                        # 過濾該 Source 的數據
                        source_df = df[df['Source'] == source].copy()
                        
                        if not source_df.empty:
                            # 計算該Source的有效轉化總金額（排除 invalid/rejected）
                            source_total_amount = 0
                            source_total_conversions = len(source_df)
                            if 'USD Sale Amount' in source_df.columns and 'Status' in source_df.columns:
                                # 過濾出有效的轉化記錄
                                valid_statuses = ['approved', 'pending', 'approved_pending']
                                valid_mask = source_df['Status'].str.lower().isin(valid_statuses)
                                valid_source_df = source_df[valid_mask]
                                source_total_amount = valid_source_df['USD Sale Amount'].sum()
                                source_total_conversions = len(valid_source_df)
                                logger.debug(f"🔍 [MONITOR] Source {source} 有效轉化金額: ${source_total_amount:,.2f} (有效記錄數: {source_total_conversions}, 總記錄數: {len(source_df)})")
                            elif 'USD Sale Amount' in source_df.columns:
                                # 如果沒有 Status 欄位，使用所有記錄
                                source_total_amount = source_df['USD Sale Amount'].sum()
                                logger.info(f"🔍 [MONITOR] Source {source} 金額: ${source_total_amount:,.2f} (記錄數: {len(source_df)})")
                            
                            # 計算該Source的無效轉化統計
                            source_invalid_count = 0
                            source_invalid_amount = 0.0
                            if 'Status' in source_df.columns:
                                # 統計該Source中狀態為invalid或rejected的記錄
                                invalid_statuses = ['invalid', 'rejected']
                                invalid_mask = source_df['Status'].str.lower().isin(invalid_statuses)
                                source_invalid_count = invalid_mask.sum()
                                if source_invalid_count > 0 and 'USD Sale Amount' in source_df.columns:
                                    source_invalid_amount = source_df.loc[invalid_mask, 'USD Sale Amount'].sum()
                            
                            # 添加匯總信息區塊（包含無效轉化警告）
                            current_row = self._add_summary_header(
                                ws, partner_name, start_date, end_date, 
                                source_total_conversions, source_total_amount, 1, source_invalid_count, source_invalid_amount, source_df
                            )
                            
                            headers = list(source_df.columns)
                            
                            # 寫入標題行
                            for col_idx, header in enumerate(headers, 1):
                                cell = ws.cell(row=current_row, column=col_idx, value=header)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = header_alignment
                            
                            # 寫入數據行
                            data_start_row = current_row + 1
                            for row_idx, (_, row) in enumerate(source_df.iterrows(), data_start_row):
                                for col_idx, value in enumerate(row, 1):
                                    # 清理Excel不支持的字符
                                    from utils.excel_character_cleaner import clean_for_excel
                                    cleaned_value = clean_for_excel(value)
                                    cell = ws.cell(row=row_idx, column=col_idx, value=cleaned_value)
                                    
                                    # 設置數據格式
                                    if isinstance(value, (int, float)):
                                        cell.alignment = number_alignment
                                        if 'USD Sale Amount' in headers[col_idx-1]:
                                            cell.number_format = '"$"#,##0.00'
                                    else:
                                        cell.alignment = data_alignment
                            
                            # 自動調整列寬
                            for column in ws.columns:
                                max_length = 0
                                column_letter = column[0].column_letter
                                for cell in column:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = min(max_length + 2, 50)
                                ws.column_dimensions[column_letter].width = adjusted_width
            else:
                # 如果 DataFrame 沒有數據，但 partner_summaries 可能有數據，需要正確計算總金額
                excel_total_amount = 0.0
                excel_total_conversions = 0
                
                if partner_summaries:
                    if partner_name.upper() == 'ALL':
                        # For 'ALL' partner, sum up all summaries
                        excel_total_amount = sum(float(s.total_amount) for s in partner_summaries)
                        excel_total_conversions = sum(s.total_records for s in partner_summaries)
                        logger.info(f"🔍 [MONITOR] Excel總表 (ALL) 空DataFrame但使用partner_summaries: ${excel_total_amount:,.2f} (記錄數: {excel_total_conversions})")
                    else:
                        # For specific partner, find the matching summary
                        target_summary = next((s for s in partner_summaries if s.partner_name == partner_name), None)
                        if target_summary:
                            excel_total_amount = float(target_summary.total_amount)
                            excel_total_conversions = target_summary.total_records
                            logger.info(f"🔍 [MONITOR] Excel總表 空DataFrame但使用partner_summaries: ${excel_total_amount:,.2f} (記錄數: {excel_total_conversions})")
                        else:
                            logger.info(f"🔍 [MONITOR] Excel總表 DataFrame為空且無partner_summaries，使用0金額")
                else:
                    logger.info(f"🔍 [MONITOR] Excel總表 DataFrame為空且partner_summaries為空，使用0金額")
                
                # 創建空的匯總信息，使用正確計算的金額
                current_row = self._add_summary_header(
                    summary_ws, partner_name, start_date, end_date, 
                    excel_total_conversions, excel_total_amount, 1, 0, 0.0, df
                )
                
                # 添加空數據提示
                summary_ws.cell(row=current_row, column=1, value="No data available for the specified criteria").font = Font(italic=True, color="666666")
            
            # 保存文件
            wb.save(file_path)
            logger.info(f"✅ Excel文件生成完成: {file_path}")
            
            return file_path
            
        except Exception as e:
            logger.error(f"❌ 生成Excel文件失败: {e}")
            raise
    
    def _clean_sheet_name(self, name: str) -> str:
        """
        清理Excel工作表名称，移除不支持的字符
        使用統一的Excel字符清理工具
        """
        from utils.excel_character_cleaner import clean_sheet_name_for_excel
        return clean_sheet_name_for_excel(name)
    
    async def _upload_to_feishu(self, excel_files: List[str]) -> Dict[str, Any]:
        """上传到飞书"""
        try:
            # 使用原有的FeishuUploader
            upload_result = self.feishu_uploader.upload_files(excel_files)
            
            if upload_result['success']:
                logger.info(f"✅ 飞书上传成功: {upload_result['success_count']} 个文件")
            else:
                logger.warning(f"⚠️ 飞书上传部分失败: 成功 {upload_result['success_count']} 个，失败 {upload_result['failed_count']} 个")
            
            return upload_result
            
        except Exception as e:
            logger.error(f"❌ 飞书上传失败: {e}")
            return {
                'success': False,
                'error': str(e),
                'success_count': 0,
                'failed_count': len(excel_files)
            }
    
    async def _send_emails(self, partner_summaries: List[PartnerSummary],
                         excel_files: List[str],
                         start_date: datetime,
                         end_date: datetime,
                         self_email: bool = False) -> Dict[str, Any]:
        """发送邮件"""
        try:
            # 准备邮件数据
            partner_data = {}
            
            for summary in partner_summaries:
                # 找到对应的Excel文件
                file_path = None
                for excel_file in excel_files:
                    if summary.partner_name in excel_file:
                        file_path = excel_file
                        break
                
                if not file_path and excel_files:
                    # 如果没有找到对应的文件，使用第一个文件
                    file_path = excel_files[0]
                
                # 获取Sources信息
                sources = []
                if hasattr(summary, 'sources'):
                    sources = summary.sources
                
                # 獲取無效轉化統計（從DataFrame的attrs中獲取）
                invalid_stats = {'invalid_count': 0, 'invalid_amount': 0.0}
                # 這裡需要從數據庫重新獲取無效轉化統計，因為partner_summaries中沒有這個信息
                try:
                    invalid_stats = await self.db.get_invalid_conversion_stats(
                        partner_name=summary.partner_name,
                        start_date=start_date,
                        end_date=end_date
                    )
                except Exception as e:
                    logger.warning(f"⚠️ 獲取 {summary.partner_name} 無效轉化統計失敗: {e}")
                
                partner_data[summary.partner_name] = {
                    'records': summary.total_records,
                    'amount_formatted': summary.amount_formatted,
                    'file_path': file_path,
                    'sources': sources,
                    'sources_count': summary.sources_count,
                    'invalid_stats': invalid_stats  # 添加無效轉化統計
                }
            
            # 发送邮件
            email_result = self.email_sender.send_partner_reports(
                partner_data,
                report_date=end_date,
                start_date=start_date,
                self_email=self_email
            )
            
            return email_result
            
        except Exception as e:
            logger.error(f"❌ 发送邮件失败: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    async def get_available_partners(self) -> List[str]:
        """获取可用的Partner列表"""
        try:
            partners = await self.db.get_available_partners()
            # 添加特殊的ALL选项
            if 'ALL' not in partners:
                partners.insert(0, 'ALL')
            return partners
        except Exception as e:
            logger.error(f"❌ 获取Partner列表失败: {e}")
            return ['ALL']
    
    async def get_partner_preview(self, partner_name: str = "ALL",
                                start_date: datetime = None,
                                end_date: datetime = None) -> Dict[str, Any]:
        """获取Partner数据预览"""
        try:
            # 设置默认日期范围
            if not end_date:
                end_date = datetime.now()
            if not start_date:
                start_date = end_date - timedelta(days=7)
            
            # 获取汇总数据
            partner_summaries = await self.db.get_partner_summary(partner_name, start_date, end_date)
            
            # 获取最近的一些转化记录作为预览
            df = await self.db.get_conversion_dataframe(partner_name, start_date, end_date)
            
            preview_data = []
            if not df.empty:
                # 取前10条记录作为预览
                preview_df = df.head(10)
                preview_data = preview_df.to_dict('records')
            
            # 計算有效轉化的總金額（排除 invalid/rejected）
            total_amount = 0
            total_records = len(df)
            if not df.empty and 'USD Sale Amount' in df.columns and 'Status' in df.columns:
                # 過濾出有效的轉化記錄
                valid_statuses = ['approved', 'pending', 'approved_pending']
                valid_mask = df['Status'].str.lower().isin(valid_statuses)
                valid_df = df[valid_mask]
                total_amount = valid_df['USD Sale Amount'].sum()
                total_records = len(valid_df)
            elif not df.empty and 'USD Sale Amount' in df.columns:
                total_amount = df['USD Sale Amount'].sum()
            
            return {
                'success': True,
                'partner_name': partner_name,
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d'),
                'total_records': total_records,
                'total_amount': total_amount,
                'partner_summaries': [summary.to_dict() for summary in partner_summaries],
                'preview_data': preview_data
            }
            
        except Exception as e:
            logger.error(f"❌ 获取Partner预览失败: {e}")
            return {
                'success': False,
                'error': str(e),
                'partner_name': partner_name
            }
    
    async def health_check(self) -> Dict[str, Any]:
        """健康检查"""
        try:
            # 检查数据库连接
            db_health = await self.db.health_check()
            
            # 检查输出目录
            output_dir_exists = self.output_dir.exists()
            output_dir_writable = os.access(self.output_dir, os.W_OK) if output_dir_exists else False
            
            return {
                'status': 'healthy' if db_health['status'] == 'healthy' else 'unhealthy',
                'database': db_health,
                'output_directory': {
                    'exists': output_dir_exists,
                    'writable': output_dir_writable,
                    'path': str(self.output_dir)
                },
                'components': {
                    'feishu_uploader': 'available',
                    'email_sender': 'available'
                },
                'timestamp': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"❌ 健康检查失败: {e}")
            return {
                'status': 'unhealthy',
                'error': str(e),
                'timestamp': datetime.now().isoformat()
            }
    
    async def cleanup(self):
        """清理资源"""
        try:
            await self.db.close_pool()
            logger.info("✅ 报表生成器资源清理完成")
        except Exception as e:
            logger.error(f"❌ 清理资源失败: {e}") 