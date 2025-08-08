#!/usr/bin/env python3
"""
邮件发送模块
负责发送转换报告邮件
"""

import os
import smtplib
import time
import socket
import zipfile
import tempfile
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
from utils.logger import print_step
import logging
from typing import List, Optional

# 获取logger实例
logger = logging.getLogger(__name__)
import config
import pandas as pd

class EmailSender:
    """邮件发送器"""
    
    def __init__(self, global_email_disabled=False):
        self.sender = config.EMAIL_SENDER
        self.default_receivers = config.EMAIL_RECEIVERS
        self.partner_email_mapping = getattr(config, 'PARTNER_EMAIL_MAPPING', config.PUB_EMAIL_MAPPING)  # 新的Partner配置
        self.partner_email_enabled = getattr(config, 'PARTNER_EMAIL_ENABLED', config.PUB_EMAIL_ENABLED)  # 新的Partner配置
        # 保持向后兼容性
        self.pub_email_mapping = self.partner_email_mapping  # 兼容性别名
        self.pub_email_enabled = self.partner_email_enabled  # 兼容性别名
        self.auto_cc_email = getattr(config, 'EMAIL_AUTO_CC', None)  # 自动抄送邮箱
        self.password = config.EMAIL_PASSWORD
        self.smtp_server = config.SMTP_SERVER
        self.smtp_port = config.SMTP_PORT
        self.enable_tls = config.EMAIL_ENABLE_TLS
        self.include_attachments = config.EMAIL_INCLUDE_ATTACHMENTS
        self.include_feishu_links = config.EMAIL_INCLUDE_FEISHU_LINKS
        self.subject_template = config.EMAIL_SUBJECT_TEMPLATE
        
        # 全局邮件禁用标志（用于--no-email参数）
        self.global_email_disabled = global_email_disabled
        
        # 新增超时和重试配置 - 增強版
        self.smtp_timeout = getattr(config, 'EMAIL_SMTP_TIMEOUT', 60)
        self.max_retries = getattr(config, 'EMAIL_MAX_RETRIES', 5)  # 增加到5次重試
        self.retry_delay = getattr(config, 'EMAIL_RETRY_DELAY', 3)  # 減少初始延遲到3秒
        self.retry_backoff = getattr(config, 'EMAIL_RETRY_BACKOFF', 1.5)  # 減少退避倍數
        self.max_retry_delay = getattr(config, 'EMAIL_MAX_RETRY_DELAY', 30)  # 最大重試延遲30秒
        
        # 🚀 階段1+階段2優化配置
        # 附件壓縮配置
        self.auto_compress_attachments = getattr(config, 'EMAIL_AUTO_COMPRESS_ATTACHMENTS', True)
        self.compress_threshold_mb = getattr(config, 'EMAIL_COMPRESS_THRESHOLD_MB', 5)
        
        # 動態超時配置
        self.dynamic_timeout_enabled = getattr(config, 'EMAIL_DYNAMIC_TIMEOUT_ENABLED', True)
        self.small_file_timeout = getattr(config, 'EMAIL_SMALL_FILE_TIMEOUT', 120)
        self.medium_file_timeout = getattr(config, 'EMAIL_MEDIUM_FILE_TIMEOUT', 300)
        self.large_file_timeout = getattr(config, 'EMAIL_LARGE_FILE_TIMEOUT', 600)
        
        # 智能降級策略配置
        self.smart_fallback_enabled = getattr(config, 'EMAIL_SMART_FALLBACK_ENABLED', True)
        self.fallback_size_threshold_mb = getattr(config, 'EMAIL_FALLBACK_SIZE_THRESHOLD_MB', 15)
        self.fallback_retry_threshold = getattr(config, 'EMAIL_FALLBACK_RETRY_THRESHOLD', 2)
        
        # 郵件模式選擇
        self.delivery_mode = getattr(config, 'EMAIL_DELIVERY_MODE', 'smart_hybrid')
        self.cloud_link_provider = getattr(config, 'EMAIL_CLOUD_LINK_PROVIDER', 'feishu')
    
    def send_partner_reports(self, partner_summary, feishu_upload_result=None, report_date=None, start_date=None, self_email=False):
        """
        按Partner分别发送转换报告邮件
        
        Args:
            partner_summary: Partner汇总数据字典，格式：
                {
                    'partner_name': {
                        'records': 数量,
                        'amount_formatted': '$金额',
                        'file_path': '文件路径',
                        'sources': ['source1', 'source2', ...],
                        'sources_count': 数量,
                        'invalid_stats': {'invalid_count': 数量, 'invalid_amount': 金额}
                    }
                }
            feishu_upload_result: 飞书上传结果
            report_date: 报告日期（结束日期，用于邮件标题和内容）
            start_date: 开始日期（用于邮件中的日期范围显示）
        
        Returns:
            dict: 发送结果汇总
        """
        print_step("Partner邮件发送", "开始按Partner分别发送转换报告邮件")
        
        # 检查全局邮件禁用标志
        if self.global_email_disabled:
            print_step("邮件发送", "⚠️ 全局邮件发送已禁用 (--no-email)，跳过所有Partner邮件发送")
            return {
                'success': True,
                'total_sent': 0,
                'total_failed': 0,
                'partner_results': {partner_name: {'success': True, 'skipped': True, 'reason': '全局邮件发送已禁用'} 
                                  for partner_name in partner_summary.keys()}
            }
        
        # 检查配置
        if self.password == "your_gmail_app_password_here":
            error_msg = "邮件密码未配置，请在config.py中设置EMAIL_PASSWORD"
            print_step("配置错误", f"❌ {error_msg}")
            return {'success': False, 'error': error_msg}
        
        # 检查SMTP连接
        if not self._test_smtp_connection():
            error_msg = "SMTP连接失败，请检查邮箱配置"
            print_step("SMTP连接", f"❌ {error_msg}")
            return {'success': False, 'error': error_msg}
        
        total_sent = 0
        total_failed = 0
        partner_results = {}
        
        for partner_name, partner_data in partner_summary.items():
            try:
                print_step("邮件发送", f"📧 正在发送 {partner_name} 的邮件...")
                
                # 检查Partner邮件是否启用
                if not self._is_partner_email_enabled(partner_name):
                    print_step("邮件发送", f"⚠️ {partner_name} 邮件发送已禁用")
                    partner_results[partner_name] = {
                        'success': True, 
                        'skipped': True, 
                        'reason': 'Partner邮件发送已禁用'
                    }
                    continue
                
                # 准备邮件数据
                email_data = self._prepare_partner_email_data(
                    partner_name, partner_data, end_date=report_date, start_date=start_date
                )
                
                # 确保email_data包含文件路径信息，供日期提取使用
                if 'file_path' not in email_data and partner_data.get('file_path'):
                    email_data['file_path'] = partner_data.get('file_path')
                    email_data['file_paths'] = [partner_data.get('file_path')]
                
                # 获取收件人列表
                recipients = self._get_partner_recipients(partner_name, self_email)
                if not recipients:
                    print_step("邮件发送", f"⚠️ {partner_name} 没有配置收件人")
                    partner_results[partner_name] = {
                        'success': False, 
                        'error': '没有配置收件人'
                    }
                    total_failed += 1
                    continue
                
                # 生成邮件正文
                email_body = self._generate_partner_email_body(partner_name, email_data, feishu_upload_result)
                
                # 获取飞书信息
                feishu_info = self._get_partner_feishu_info(partner_name, feishu_upload_result)
                
                # 发送邮件（使用重试机制）
                send_result = self._send_single_partner_email(
                    partner_name=partner_name,
                    email_data=email_data,
                    file_paths=[email_data['file_path']] if email_data['file_path'] else [],
                    receivers=recipients,
                    feishu_info=feishu_info,
                    report_date=report_date
                )
                
                if send_result['success']:
                    print_step("邮件发送", f"✅ {partner_name} 邮件发送成功 (尝试 {send_result.get('attempts', 1)} 次)")
                    partner_results[partner_name] = {'success': True, 'attempts': send_result.get('attempts', 1)}
                    total_sent += 1
                else:
                    print_step("邮件发送", f"❌ {partner_name} 邮件发送失败: {send_result.get('error', '发送失败')}")
                    partner_results[partner_name] = {'success': False, 'error': send_result.get('error', '发送失败')}
                    total_failed += 1
                
            except Exception as e:
                print_step("邮件发送", f"❌ {partner_name} 邮件发送异常: {str(e)}")
                partner_results[partner_name] = {'success': False, 'error': str(e)}
                total_failed += 1
        
        result = {
            'success': total_failed == 0,
            'total_sent': total_sent,
            'total_failed': total_failed,
            'partner_results': partner_results
        }
        
        if result['success']:
            print_step("邮件发送", f"✅ 所有邮件发送成功: {total_sent} 封")
        else:
            print_step("邮件发送", f"⚠️ 邮件发送完成: 成功 {total_sent} 封，失败 {total_failed} 封")
        
        return result
    
    def send_report_email(self, report_data, file_paths=None, feishu_upload_result=None):
        """
        发送转换报告邮件（兼容性保留，建议使用send_pub_reports）
        
        Args:
            report_data: 报告数据字典
            file_paths: Excel文件路径列表
            feishu_upload_result: 飞书上传结果
        
        Returns:
            dict: 发送结果
        """
        print_step("邮件发送", "开始发送转换报告邮件")
        
        # 检查配置
        if self.password == "your_gmail_app_password_here":
            error_msg = "邮件密码未配置，请在config.py中设置EMAIL_PASSWORD"
            print_step("配置错误", f"❌ {error_msg}")
            return {'success': False, 'error': error_msg}
        
        try:
            # 创建邮件对象
            msg = self._create_email_message(report_data, file_paths, feishu_upload_result, self.default_receivers)
            
            # 使用重试机制发送邮件
            send_result = self._send_email_with_retry(msg, self.default_receivers, "通用邮件发送")
            
            if send_result['success']:
                print_step("邮件发送成功", f"✅ 邮件已发送给 {', '.join(self.default_receivers)} (尝试 {send_result['attempts']} 次)")
                return {
                    'success': True,
                    'recipients': self.default_receivers,
                    'attachments_count': len(file_paths) if file_paths and self.include_attachments else 0,
                    'attempts': send_result['attempts']
                }
            else:
                print_step("邮件发送失败", f"❌ {send_result['error']}")
                return {'success': False, 'error': send_result['error']}
            
        except Exception as e:
            error_msg = f"邮件准备失败: {str(e)}"
            print_step("邮件发送失败", f"❌ {error_msg}")
            return {'success': False, 'error': error_msg}
    
    def _send_single_partner_email(self, partner_name, email_data, file_paths, receivers, feishu_info, report_date=None):
        """发送单个Partner的邮件"""
        try:
            # 添加抄送邮箱（从配置读取）
            cc_email = self.auto_cc_email
            all_recipients = receivers.copy()
            if cc_email and cc_email not in receivers:
                all_recipients.append(cc_email)
            
            # 创建邮件对象
            msg = self._create_partner_email_message(partner_name, email_data, file_paths, receivers, feishu_info, report_date, cc_email)
            
            # 使用重试机制发送邮件
            operation_name = f"Partner邮件-{partner_name}"
            send_result = self._send_email_with_retry(msg, all_recipients, operation_name)
            
            if send_result['success']:
                cc_info = f" (抄送: {cc_email})" if cc_email else ""
                print_step(f"Partner邮件-{partner_name}", f"✅ 邮件已发送给 {', '.join(receivers)}{cc_info} (尝试 {send_result['attempts']} 次)")
                return {
                    'success': True,
                    'recipients': receivers,
                    'cc_recipients': [cc_email] if cc_email else [],
                    'attachments_count': len(file_paths) if file_paths and self.include_attachments else 0,
                    'attempts': send_result['attempts']
                }
            else:
                print_step(f"Partner邮件-{partner_name}", f"❌ {send_result['error']}")
                return {'success': False, 'error': send_result['error']}
            
        except Exception as e:
            error_msg = f"邮件准备失败: {str(e)}"
            print_step(f"Partner邮件-{partner_name}", f"❌ {error_msg}")
            return {'success': False, 'error': error_msg}
    
    def _prepare_partner_email_data(self, partner_name, partner_data, end_date=None, start_date=None):
        """准备Partner邮件数据"""
        file_path = partner_data.get('file_path')
        
        # 🎯 修復：優先使用傳入的日期參數，只有在缺失時才從文件名提取
        if start_date is not None and end_date is not None:
            # 使用傳入的日期參數
            print_step("邮件数据准备", f"📅 使用傳入的日期範圍: {start_date} to {end_date}")
            
            # 如果是 datetime 對象，轉換為字符串
            if hasattr(end_date, 'strftime'):
                end_date = end_date.strftime("%Y-%m-%d")
            if hasattr(start_date, 'strftime'):
                start_date = start_date.strftime("%Y-%m-%d")
        else:
            # 備用方案：從文件名中提取報表日期
            extracted_report_date = None
            if file_path:
                extracted_report_date = self._extract_report_date_from_filename(file_path)
                if extracted_report_date:
                    print_step("邮件数据准备", f"📅 使用文件名中的报表日期: {extracted_report_date}")
                    end_date = extracted_report_date
                    start_date = extracted_report_date
                else:
                    print_step("邮件数据准备", f"⚠️ 无法从文件名提取日期，使用當前日期")
            
            # 如果仍然没有日期，使用當前日期
            if end_date is None:
                end_date = datetime.now().strftime("%Y-%m-%d")
            elif hasattr(end_date, 'strftime'):
                end_date = end_date.strftime("%Y-%m-%d")
                
            if start_date is None:
                start_date = end_date
            elif hasattr(start_date, 'strftime'):
                start_date = start_date.strftime("%Y-%m-%d")
        
        # 从Excel文件中计算真实的销售总额
        real_total_amount = self._calculate_sales_amount_from_excel(file_path)
        
        # 计算Sources统计信息
        sources_statistics = self._calculate_sources_statistics_from_excel(file_path)
        
        # 使用统一Summary格式化器生成状态统计
        try:
            # 读取Excel文件数据用于统一Summary生成
            df = None
            if file_path and os.path.exists(file_path):
                try:
                    # 尝试读取Excel文件数据
                    df = pd.read_excel(file_path)
                    
                    # 如果第一行不是标准列名，尝试跳过几行（擴展範圍到15行）
                    if 'Status' not in df.columns:
                        for skip_rows in range(1, 15):  # 擴展到15行，因為Reporter生成的Excel有較多標題行
                            try:
                                test_df = pd.read_excel(file_path, skiprows=skip_rows)
                                if 'Status' in test_df.columns:
                                    df = test_df
                                    break
                            except:
                                continue
                    
                    # 檢查Status列的值分佈
                    if df is not None and 'Status' in df.columns:
                        status_counts = df['Status'].value_counts()
                    else:
                        logger.warning(f"   ⚠️ 仍然沒有找到Status列")
                        
                except Exception as e:
                    logger.error(f"⚠️ 读取Excel文件失败: {e}")
                    df = None
            
            # 生成统一Summary
            from shared.utils.summary_formatter import generate_unified_summary
            summary = generate_unified_summary(
                partner_name=partner_name,
                start_date=start_date,
                end_date=end_date,
                df=df,
                total_records=partner_data.get('records', 0),
                total_amount=float(real_total_amount.replace('$', '').replace(',', '')) if isinstance(real_total_amount, str) else real_total_amount,
                sources=[stat.get('source_name', '') for stat in sources_statistics]
            )
            
            # 使用统一Summary的数据
            status_statistics = {
                'total_all_conversions': summary.get('total_all_conversions', 0),
                'pending_approved_count': summary.get('pending_approved_count', 0),
                'pending_approved_amount': summary.get('pending_approved_amount', '$0.00'),
                'pending_approved_amount_numeric': summary.get('pending_approved_amount_numeric', 0.0),
                'invalid_rejected_count': summary.get('invalid_rejected_count', 0),
                'invalid_rejected_amount': summary.get('invalid_rejected_amount', '$0.00'),
                'invalid_rejected_amount_numeric': summary.get('invalid_rejected_amount_numeric', 0.0)
            }
            
        except Exception as e:
            print(f"❌ 生成统一Summary失败: {e}")
            # 回退到旧的状态统计方法
            status_statistics = self._calculate_status_statistics_from_excel(file_path)
            # 确保包含numeric字段
            status_statistics.update({
                'pending_approved_amount_numeric': 0.0,
                'invalid_rejected_amount_numeric': 0.0
            })
        
        # 獲取無效轉化統計（如果存在）
        invalid_stats = partner_data.get('invalid_stats', {'invalid_count': 0, 'invalid_amount': 0.0})
        
        return {
            'partner_name': partner_name,
            'total_records': partner_data.get('records', 0),
            'total_amount': real_total_amount,
            'start_date': start_date,
            'end_date': end_date,
            'report_date': end_date,
            'main_file': os.path.basename(file_path) if file_path else '',
            'file_path': file_path,  # 添加完整文件路径，供ByteC邮件模板使用
            'sources': partner_data.get('sources', []),
            'sources_count': partner_data.get('sources_count', 0),
            'sources_statistics': sources_statistics,
            'invalid_stats': invalid_stats,  # 添加無效轉化統計
            # 添加详细状态统计（同步Excel Summary）
            'total_all_conversions': status_statistics.get('total_all_conversions', 0),
            'pending_approved_count': status_statistics.get('pending_approved_count', 0),
            'pending_approved_amount': status_statistics.get('pending_approved_amount', '$0.00'),
            'pending_approved_amount_numeric': status_statistics.get('pending_approved_amount_numeric', 0.0),
            'invalid_rejected_count': status_statistics.get('invalid_rejected_count', 0),
            'invalid_rejected_amount': status_statistics.get('invalid_rejected_amount', '$0.00'),
            'invalid_rejected_amount_numeric': status_statistics.get('invalid_rejected_amount_numeric', 0.0)
        }
    
    def _calculate_sales_amount_from_excel(self, file_path):
        """从Excel文件中计算Sale Amount总额（包含所有sheets）"""
        try:
            if not file_path or not os.path.exists(file_path):
                import traceback
                tb = traceback.format_stack()
                print_step("金额计算", f"⚠️ 文件不存在: {file_path}\n调用栈:\n{''.join(tb)}")
                return '$0.00'
            
            import openpyxl
            
            # 使用openpyxl读取所有sheets
            wb = openpyxl.load_workbook(file_path, read_only=True)
            total_amount = 0.0
            sheet_details = []
            
            logger.info(f"📊 金额计算: 正在计算 {os.path.basename(file_path)} 的销售总额（包含所有sheets）...")
            
            for sheet_name in wb.sheetnames:
                try:
                    # 新的逻辑：从Excel文件计算总金额
                    # 1. 如果只有Partner主sheet，使用主sheet的金额
                    # 2. 如果有多个sheets（主sheet + source sheets），使用主sheet避免重复计算
                    # 3. 优先使用Partner主sheet，如果没有则累加所有sheets
                    
                    # 判断是否为Partner主sheet：sheet名称与partner名称匹配
                    partner_name = os.path.basename(file_path).split('_')[0]  # 从文件名提取partner名称
                    is_partner_main_sheet = (sheet_name == partner_name)
                    
                    # 读取该sheet的数据
                    # 首先嘗試正常讀取，如果有Summary信息則會在後面處理
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    # 如果第一行不是標準的列名，可能是新格式的Summary信息
                    # 嘗試找到真正的數據開始行
                    if 'USD Sale Amount' not in df.columns:
                        # 嘗試不同的skiprows值來找到數據
                        for skip_rows in range(1, 15):
                            try:
                                test_df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=skip_rows)
                                if 'USD Sale Amount' in test_df.columns:
                                    df = test_df
                                    break
                            except:
                                continue
                    
                    # 支持多种可能的销售金额列名
                    sales_amount_col = None
                    possible_col_names = ['USD Sale Amount', 'sale_amount', 'Sale Amount', 'sales_amount', 'SALE_AMOUNT']
                    
                    for col_name in possible_col_names:
                        if col_name in df.columns:
                            sales_amount_col = col_name
                            break
                    
                    if sales_amount_col and len(df) > 0:
                        sheet_total = df[sales_amount_col].sum()
                        
                        # 如果是Partner主sheet，直接使用其金额作为总金额并停止计算
                        if is_partner_main_sheet:
                            total_amount = sheet_total
                            sheet_details.append(f"  - {sheet_name}: ${sheet_total:.2f} (Partner主sheet)")
                            logger.info(f"📋 金额计算: 使用Partner主sheet '{sheet_name}': ${sheet_total:.2f} ({len(df)} 条记录)")
                            break  # 找到Partner主sheet就停止
                        else:
                            # 如果不是Partner主sheet，累加到总金额
                            total_amount += sheet_total
                            sheet_details.append(f"  - {sheet_name}: ${sheet_total:.2f} (Source sheet)")
                            logger.info(f"📋 金额计算: 累加Source sheet '{sheet_name}': ${sheet_total:.2f} ({len(df)} 条记录)")
                    else:
                        if 'USD Sale Amount' not in df.columns:
                            logger.info(f"⚠️ 金额计算: Sheet '{sheet_name}': 未找到销售金额列，可用列: {list(df.columns[:5])}")
                        else:
                            logger.info(f"⚠️ 金额计算: Sheet '{sheet_name}': 数据为空")
                        sheet_details.append(f"  - {sheet_name}: $0.00 (无数据)")
                    
                except Exception as e:
                    logger.info(f"❌ 金额计算: 读取Sheet '{sheet_name}' 失败: {str(e)}")
                    sheet_details.append(f"  - {sheet_name}: 读取失败")
            
            wb.close()
            
            formatted_amount = f"${total_amount:,.2f}"
            logger.info(f"💰 金额计算: {os.path.basename(file_path)} 总销售额: {formatted_amount}")
            logger.info(f"金额详情: 各Sheet明细:\n" + "\n".join(sheet_details))
            
            return formatted_amount
                
        except Exception as e:
            logger.info(f"❌ 金额计算: 计算金额失败 {os.path.basename(file_path)}: {str(e)}")
            return '$0.00'
    
    def _calculate_sources_statistics_from_excel(self, file_path):
        """从Excel文件中计算各Sources的统计信息（直接使用原始Source值）"""
        try:
            if not file_path or not os.path.exists(file_path):
                return []
            
            import openpyxl
            
            # 使用openpyxl读取多个sheets
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sources_stats = []
            
            print_step("Sources统计", f"📊 正在计算 {os.path.basename(file_path)} 的Sources统计...")
            
            # 从文件名提取partner名称
            partner_name = os.path.basename(file_path).split('_')[0]
            
            for sheet_name in wb.sheetnames:
                try:
                    # 判断是否为Partner主sheet
                    is_partner_main_sheet = (sheet_name == partner_name)
                    
                    # 跳过Partner主sheet，只处理Source sheets
                    if is_partner_main_sheet:
                        # print_step("Sources统计", f"📋 跳过Sheet '{sheet_name}': Partner主sheet")
                        continue
                    
                    # 这是Source sheet，进行统计
                    # print_step("Sources统计", f"📋 处理Source sheet '{sheet_name}'")
                    
                    ws = wb[sheet_name]
                    
                    # 读取该sheet的数据来计算销售金额
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    # 如果第一行不是標準的列名，可能是新格式的Summary信息
                    # 嘗試找到真正的數據開始行
                    if 'USD Sale Amount' not in df.columns:
                        # 嘗試不同的skiprows值來找到數據
                        for skip_rows in range(1, 15):
                            try:
                                test_df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=skip_rows)
                                if 'USD Sale Amount' in test_df.columns:
                                    df = test_df
                                    break
                            except:
                                continue
                    
                    # 支持多种可能的销售金额列名
                    sales_amount_col = None
                    possible_col_names = ['USD Sale Amount', 'sale_amount', 'Sale Amount', 'sales_amount', 'SALE_AMOUNT']
                    
                    for col_name in possible_col_names:
                        if col_name in df.columns:
                            sales_amount_col = col_name
                            break
                    
                    if sales_amount_col and len(df) > 0:
                        # 处理格式化的美元金额字符串（如"$123.45"）
                        def parse_currency(value):
                            """解析货币字符串，返回数值"""
                            if pd.isna(value):
                                return 0.0
                            if isinstance(value, str):
                                # 移除美元符号、逗号和其他非数字字符
                                cleaned_value = value.replace('$', '').replace(',', '').strip()
                                try:
                                    return float(cleaned_value)
                                except ValueError:
                                    return 0.0
                            return float(value) if value else 0.0
                        
                        sales_amount = df[sales_amount_col].apply(parse_currency).sum()
                        formatted_amount = f"${sales_amount:,.2f}"
                        row_count = len(df)
                        # print_step("Sources统计", f"📋 Sheet '{sheet_name}': {formatted_amount} ({row_count} 条记录，使用列'{sales_amount_col}')")
                    else:
                        formatted_amount = '$0.00'
                        row_count = len(df) if not df.empty else 0
                        # print_step("Sources统计", f"⚠️ Sheet '{sheet_name}': 无销售金额列或无数据")
                    
                    # 直接使用sheet_name作为source_name（就是原始的aff_sub1值）
                    sources_stats.append({
                        'source_name': sheet_name,
                        'records': row_count,
                        'sales_amount': formatted_amount
                    })
                
                except Exception as e:
                    # print_step("Sources统计", f"⚠️ 处理Sheet '{sheet_name}' 失败: {str(e)}")
                    sources_stats.append({
                        'source_name': sheet_name,
                        'records': 0,
                        'sales_amount': '$0.00'
                    })
            
            wb.close()
            print_step("Sources统计", f"✅ 成功计算 {len(sources_stats)} 个Sources统计")
            return sources_stats
            
        except Exception as e:
            print_step("Sources统计", f"❌ 计算Sources统计失败 {os.path.basename(file_path)}: {str(e)}")
            return []
    
    def _calculate_status_statistics_from_excel(self, file_path):
        """从Excel文件中计算详细状态统计信息（同步Excel Summary格式）"""
        try:
            if not file_path or not os.path.exists(file_path):
                import traceback
                tb = traceback.format_stack()
                print_step("状态统计", f"⚠️ 文件不存在: {file_path}\n调用栈:\n{''.join(tb)}")
                return {
                    'total_all_conversions': 0, 
                    'pending_approved_count': 0, 
                    'pending_approved_amount': '$0.00', 
                    'invalid_rejected_count': 0, 
                    'invalid_rejected_amount': '$0.00'
                }
            
            import pandas as pd
            
            print_step("状态统计", f"📊 正在计算 {os.path.basename(file_path)} 的状态统计...")
            
            # 获取partner名称（从文件名提取）
            partner_name = os.path.basename(file_path).split('_')[0]
            
            # 优先读取Partner主sheet的数据
            try:
                # 首先尝试读取Partner主sheet
                df = pd.read_excel(file_path, sheet_name=partner_name)
                
                # 如果第一行不是标准列名，尝试跳过Summary行
                if 'Status' not in df.columns:
                    for skip_rows in range(1, 15):
                        try:
                            test_df = pd.read_excel(file_path, sheet_name=partner_name, skiprows=skip_rows)
                            if 'Status' in test_df.columns:
                                df = test_df
                                print_step("状态统计", f"📋 找到数据起始行: {skip_rows + 1}")
                                break
                        except:
                            continue
                
                # 检查是否找到Status列
                if 'Status' not in df.columns:
                    print_step("状态统计", f"⚠️ 未找到Status列，可用列: {list(df.columns[:5])}")
                    return {
                        'total_all_conversions': 0, 
                        'pending_approved_count': 0, 
                        'pending_approved_amount': '$0.00', 
                        'invalid_rejected_count': 0, 
                        'invalid_rejected_amount': '$0.00'
                    }
                
                # 计算状态统计
                total_all_conversions = len(df)
                
                # 计算各状态的转化数 - 修复状态识别逻辑
                # 根据运行日志，使用更准确的状态分类
                
                # 检查Status列的实际值
                unique_statuses = df['Status'].value_counts()
                print_step("状态统计", f"📋 发现状态类型: {dict(unique_statuses)}")
                
                # 无效状态：包含invalid, rejected, cancelled, failed等
                invalid_keywords = ['invalid', 'rejected', 'cancelled', 'canceled', 'failed', 'decline']
                invalid_rejected_mask = df['Status'].str.lower().str.contains('|'.join(invalid_keywords), na=False)
                
                # 有效状态：除了无效状态外的所有其他状态
                pending_approved_mask = ~invalid_rejected_mask
                
                pending_approved_count = pending_approved_mask.sum()
                invalid_rejected_count = invalid_rejected_mask.sum()
                
                # 计算各状态的金额 (如果有 USD Sale Amount 列)
                pending_approved_amount = 0.0
                invalid_rejected_amount = 0.0
                if 'USD Sale Amount' in df.columns:
                    # 处理可能的格式化字符串金额
                    def parse_currency(value):
                        """解析货币字符串，返回数值"""
                        if pd.isna(value):
                            return 0.0
                        if isinstance(value, str):
                            # 移除美元符号、逗号和其他非数字字符
                            cleaned_value = value.replace('$', '').replace(',', '').strip()
                            try:
                                return float(cleaned_value)
                            except ValueError:
                                return 0.0
                        return float(value) if value else 0.0
                    
                    # 计算有效转化金额
                    valid_amounts = df[pending_approved_mask]['USD Sale Amount'].apply(parse_currency)
                    pending_approved_amount = valid_amounts.sum()
                    
                    # 计算无效转化金额
                    invalid_amounts = df[invalid_rejected_mask]['USD Sale Amount'].apply(parse_currency)
                    invalid_rejected_amount = invalid_amounts.sum()
                
                print_step("状态统计", f"📊 统计结果: 总计 {total_all_conversions}, 有效 {pending_approved_count}, 无效 {invalid_rejected_count}")
                print_step("状态统计", f"💰 金额统计: 有效 ${pending_approved_amount:,.2f}, 无效 ${invalid_rejected_amount:,.2f}")
                
                return {
                    'total_all_conversions': total_all_conversions,
                    'pending_approved_count': pending_approved_count,
                    'pending_approved_amount': f"${pending_approved_amount:,.2f}",
                    'invalid_rejected_count': invalid_rejected_count,
                    'invalid_rejected_amount': f"${invalid_rejected_amount:,.2f}"
                }
                
            except Exception as e:
                print_step("状态统计", f"❌ 读取Partner主sheet失败: {str(e)}")
                return {
                    'total_all_conversions': 0, 
                    'pending_approved_count': 0, 
                    'pending_approved_amount': '$0.00', 
                    'invalid_rejected_count': 0, 
                    'invalid_rejected_amount': '$0.00'
                }
            
        except Exception as e:
            print_step("状态统计", f"❌ 计算状态统计失败 {os.path.basename(file_path) if file_path else 'None'}: {str(e)}")
            return {
                'total_all_conversions': 0, 
                'pending_approved_count': 0, 
                'pending_approved_amount': '$0.00', 
                'invalid_rejected_count': 0, 
                'invalid_rejected_amount': '$0.00'
            }
    
    def _get_partner_feishu_info(self, partner_name, feishu_upload_result):
        """获取该Partner的飞书文件信息"""
        if not feishu_upload_result or not feishu_upload_result.get('uploaded_files'):
            return None
        
        # 查找该Partner对应的飞书文件
        for file_info in feishu_upload_result['uploaded_files']:
            filename = file_info.get('filename', '')
            if filename.startswith(partner_name):
                return {
                    'success': True,
                    'uploaded_files': [file_info]
                }
        
        return None
    
    def _create_partner_email_message(self, partner_name, email_data, file_paths, receivers, feishu_info, report_date=None, cc_email=None):
        """创建Partner邮件消息"""
        # 优先使用从文件名中提取的报表日期，而不是当前日期
        if file_paths and len(file_paths) > 0:
            # 从文件名中提取报表日期
            extracted_report_date = self._extract_report_date_from_filename(file_paths[0])
            if extracted_report_date:
                report_date = extracted_report_date
                print_step("邮件日期", f"📅 使用文件名中的报表日期: {report_date}")
            else:
                print_step("邮件日期", f"⚠️ 无法从文件名提取日期，使用传入的report_date: {report_date}")
        
        # 生成邮件主题 - 使用报表日期而非当前日期
        start_date = email_data.get('start_date', report_date or datetime.now().strftime("%Y-%m-%d"))
        end_date = email_data.get('end_date', report_date or datetime.now().strftime("%Y-%m-%d"))
        
        # 使用檔名作為郵件主題
        if file_paths and len(file_paths) > 0:
            filename = os.path.basename(file_paths[0])
            subject = filename
        else:
            # 備用方案：統一郵件標題格式：Partner名 + Conversion Report + 日期範圍
            subject = f"{partner_name} Conversion Report - {start_date}_to_{end_date}"
        
        # 创建邮件对象
        msg = MIMEMultipart()
        msg['From'] = self.sender
        msg['To'] = ", ".join(receivers)
        if cc_email:
            msg['Cc'] = cc_email
        msg['Subject'] = subject
        
        # 生成邮件正文
        body = self._generate_partner_email_body(partner_name, email_data, feishu_info)
        msg.attach(MIMEText(body, 'html'))
        
        # 🚀 根據配置的郵件模式處理附件
        self._handle_attachments_by_mode(msg, file_paths, feishu_info)
        
        return msg

    def _extract_report_date_from_filename(self, file_path: str) -> Optional[str]:
        """从文件名中提取报表日期"""
        try:
            import re
            filename = os.path.basename(file_path)
            
            # 匹配8位数字日期格式（YYYYMMDD）
            match = re.search(r'(\d{8})', filename)
            if match:
                date_str = match.group(1)
                # 格式化为 YYYY-MM-DD
                formatted_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
                return formatted_date
            
            # 如果没有找到8位数字，尝试其他日期格式
            # 匹配 YYYY-MM-DD 格式
            match = re.search(r'(\d{4}-\d{2}-\d{2})', filename)
            if match:
                return match.group(1)
            
            return None
            
        except Exception as e:
            print_step("日期提取", f"⚠️ 从文件名提取日期失败: {e}")
            return None

    def _handle_attachments_by_mode(self, msg, file_paths, feishu_info):
        """根據配置的郵件模式處理附件"""
        if not file_paths or not self.include_attachments:
            return
        
        print_step("郵件模式", f"📧 使用郵件模式: {self.delivery_mode}")
        
        if self.delivery_mode == "cloud_link":
            # 雲端鏈接模式：總是不發送附件，添加飛書鏈接說明
            self._add_cloud_link_notice_to_email(msg, file_paths, feishu_info)
            
        elif self.delivery_mode == "attachment":
            # 附件模式：總是嘗試發送附件（使用壓縮優化）
            self._attach_files_with_compression(msg, file_paths)
            
        elif self.delivery_mode == "smart_hybrid":
            # 智能混合模式：根據文件大小決定策略
            self._smart_attachment_handling(msg, file_paths, feishu_info)
        
        else:
            # 默認行為（向後兼容）
            self._attach_files_with_compression(msg, file_paths)

    def _attach_files_with_compression(self, msg, file_paths):
        """使用壓縮優化發送附件"""
        for file_path in file_paths:
            if os.path.exists(file_path):
                self._attach_file(msg, file_path)

    def _smart_attachment_handling(self, msg, file_paths, feishu_info):
        """智能附件處理：根據文件大小決定策略"""
        large_files = []
        
        for file_path in file_paths:
            if os.path.exists(file_path):
                file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
                
                # 檢查是否應該跳過大文件
                if file_size_mb > self.fallback_size_threshold_mb:
                    large_files.append((file_path, file_size_mb))
                    print_step("智能附件", f"⚠️ 跳過大附件: {os.path.basename(file_path)} ({file_size_mb:.1f}MB) - 超過閾值 {self.fallback_size_threshold_mb}MB")
                else:
                    self._attach_file(msg, file_path)
        
        # 如果有大文件被跳過，添加雲端說明
        if large_files:
            self._add_large_file_notice_to_email(msg, large_files, feishu_info)

    def _add_cloud_link_notice_to_email(self, msg, file_paths, feishu_info):
        """添加雲端鏈接說明到郵件中"""
        body_part = None
        for part in msg.get_payload():
            if part.get_content_type() == "text/html":
                body_part = part
                break
        
        if body_part:
            current_body = body_part.get_payload()
            
            # 生成飛書鏈接信息
            feishu_links_html = ""
            if feishu_info and feishu_info.get('uploaded_files'):
                feishu_links_html = "<h4 style='color: #0c5460; margin: 15px 0 10px 0;'>📄 飛書文件鏈接：</h4><ul style='color: #0c5460; margin-bottom: 15px;'>"
                for file_info in feishu_info['uploaded_files']:
                    filename = file_info.get('filename', 'Unknown')
                    file_id = file_info.get('file_id', '')
                    if file_id:
                        # 生成飛書文件鏈接
                        feishu_link = f"https://www.feishu.cn/sheets/{file_id}"
                        feishu_links_html += f"<li><a href='{feishu_link}' style='color: #007bff; text-decoration: none;'>{filename}</a></li>"
                    else:
                        feishu_links_html += f"<li>{filename} (文件ID: {file_id})</li>"
                feishu_links_html += "</ul>"
            
            cloud_notice = f"""
            <div style='margin-top: 30px; padding: 20px; background-color: #e8f4fd; border: 1px solid #bee5eb; border-radius: 8px;'>
                <h3 style='color: #0c5460; margin-top: 0;'>☁️ 雲端文件訪問</h3>
                <p style='color: #0c5460; margin-bottom: 10px;'>根據配置，本郵件採用雲端鏈接模式，文件已上傳至飛書存儲：</p>
                {feishu_links_html}
                <p style='color: #0c5460; margin-bottom: 10px;'>其他訪問方式：</p>
                <ol style='color: #0c5460; margin-bottom: 0;'>
                    <li>點擊上方飛書鏈接直接訪問文件</li>
                    <li>登入飛書工作台查看最新上傳的Excel報告</li>
                    <li>聯繫系統管理員獲取文件訪問權限</li>
                </ol>
                <p style='color: #6c757d; font-size: 0.9em; margin-top: 15px; margin-bottom: 0;'>
                    💡 優點：快速發送、無附件大小限制、支援線上協作、即時同步更新
                </p>
            </div>
            """
            updated_body = current_body.replace("</body>", cloud_notice + "</body>")
            body_part.set_payload(updated_body)

    def _add_large_file_notice_to_email(self, msg, large_files, feishu_info):
        """添加大文件跳過說明到郵件中"""
        body_part = None
        for part in msg.get_payload():
            if part.get_content_type() == "text/html":
                body_part = part
                break
        
        if body_part:
            current_body = body_part.get_payload()
            large_files_notice = "<div style='margin-top: 20px; padding: 15px; background-color: #fff3cd; border: 1px solid #ffeaa7; border-radius: 5px;'>"
            large_files_notice += "<h4 style='color: #856404; margin: 0 0 10px 0;'>📎 大文件說明</h4>"
            large_files_notice += "<p style='margin: 0; color: #856404;'>以下文件因大小限制未作為附件發送：</p><ul style='margin: 10px 0 0 20px; color: #856404;'>"
            for file_path, size_mb in large_files:
                large_files_notice += f"<li>{os.path.basename(file_path)} ({size_mb:.1f}MB)</li>"
            large_files_notice += "</ul><p style='margin: 10px 0 0 0; color: #856404;'>請查看飛書上傳結果或聯繫系統管理員獲取完整報告文件。</p></div>"
            
            updated_body = current_body.replace("</body>", large_files_notice + "</body>")
            body_part.set_payload(updated_body)
    
    def _create_email_message(self, report_data, file_paths=None, feishu_upload_result=None, receivers=None):
        """创建邮件消息（兼容性保留）"""
        if receivers is None:
            receivers = self.default_receivers
            
        # 生成邮件主题
        today = datetime.now().strftime("%Y-%m-%d")
        subject = self.subject_template.format(date=today)
        
        # 创建邮件对象
        msg = MIMEMultipart()
        msg['From'] = self.sender
        msg['To'] = ", ".join(receivers)
        msg['Subject'] = subject
        
        # 生成邮件正文
        body = self._generate_email_body(report_data, feishu_upload_result)
        msg.attach(MIMEText(body, 'html'))
        
        # 智能附件處理：Gmail對大附件有嚴格限制，自動跳過大文件避免發送失敗
        large_files = []
        if file_paths and self.include_attachments:
            for file_path in file_paths:
                if os.path.exists(file_path):
                    file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
                    # Gmail建議附件不超過25MB，但實際使用中10MB以上容易失敗
                    if file_size_mb > 8:  # 設置更保守的8MB限制確保郵件能發送成功
                        large_files.append((file_path, file_size_mb))
                        logger.info(f"⚠️ 跳過大附件: {os.path.basename(file_path)} ({file_size_mb:.1f}MB) - Gmail SMTP限制")
                    else:
                        self._attach_file(msg, file_path)
        
        # 如果有大文件被跳過，在郵件中說明（兼容性方法）
        if large_files:
            logger.info(f"📧 兼容性郵件跳過了 {len(large_files)} 個大附件以確保發送成功")
        
        return msg
    
    def _load_html_template(self, template_name):
        """加载HTML模板文件"""
        try:
            # 獲取項目根目錄路徑
            current_dir = os.path.dirname(os.path.abspath(__file__))
            project_root = os.path.join(current_dir, '..', '..')
            project_root = os.path.normpath(project_root)
            template_path = os.path.join(project_root, 'templates', template_name)
            
            with open(template_path, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            print_step("模板加载", f"❌ 加载模板失败 {template_name}: {str(e)}")
            return None
    
    def _generate_partner_email_body(self, partner_name, email_data, feishu_info):
        """生成Partner专用邮件正文"""
        # 检查是否为ByteC Partner，使用专用模板
        if self._is_bytec_partner(partner_name):
            return self._generate_bytec_email_body(partner_name, email_data, feishu_info)
        
        # 从email_data获取Partner专用数据
        total_records = email_data.get('total_records', 0)
        total_amount = email_data.get('total_amount', '$0.00')
        start_date = email_data.get('start_date', '')
        end_date = email_data.get('end_date', '')
        report_date = email_data.get('report_date', end_date)  # 使用 end_date 作為默認值
        
        # 🎯 修復：僅在缺少 start_date/end_date 時才使用文件名中的日期
        # 避免覆蓋已經正確設置的日期參數
        if not start_date or not end_date:
            if 'file_paths' in email_data and email_data['file_paths']:
                extracted_report_date = self._extract_report_date_from_filename(email_data['file_paths'][0])
                if extracted_report_date:
                    if not start_date:
                        start_date = extracted_report_date
                    if not end_date:
                        end_date = extracted_report_date
                    report_date = extracted_report_date
                    print_step("邮件正文日期", f"📅 使用文件名中的报表日期補充缺失的日期: {report_date}")
        
        # 確保 Date Range 使用正確的日期參數
        final_start_date = start_date
        final_end_date = end_date
        print_step("邮件正文日期", f"📅 最終使用的Date Range: {final_start_date} to {final_end_date}")
        
        # 使用報告日期而不是當前日期
        completion_time = f"{report_date} {datetime.now().strftime('%H:%M:%S')}"
        main_file = email_data.get('main_file', f'{partner_name}_ConversionReport_{report_date}.xlsx')
        sources_statistics = email_data.get('sources_statistics', [])
        
        # 獲取無效轉化統計
        invalid_stats = email_data.get('invalid_stats', {'invalid_count': 0, 'invalid_amount': 0.0})
        invalid_count = invalid_stats.get('invalid_count', 0)
        invalid_amount = invalid_stats.get('invalid_amount', 0.0)
        
        # 加载HTML模板
        template = self._load_html_template('email_template.html')
        if not template:
            # 如果模板加载失败，使用备用的简单HTML
            logger.warning(f"⚠️ 模板加载失败，使用备用HTML")
            return self._generate_fallback_email_body(partner_name, email_data, feishu_info)
        
        logger.info(f"✅ 成功加载邮件模板")
        
        # 准备飞书链接部分
        feishu_section = ""
        if feishu_info and feishu_info.get('success'):
            feishu_section = f"""
            <h3 style="color: #1f4e79;">☁️ Feishu File Links:</h3>
            <ul>
            """
            for file_info in feishu_info.get('uploaded_files', []):
                if file_info.get('url'):
                    feishu_section += f'<li><a href="{file_info["url"]}">{file_info["filename"]}</a></li>'
            feishu_section += "</ul>"
        
        # 准备Sources列表 - 使用sources_statistics生成
        sources_list = self._generate_sources_list(sources_statistics)
        if not sources_list or sources_list == "无":
            sources_list = "N/A"
        
        # 生成Sources统计HTML
        sources_statistics_html = self._generate_sources_statistics_html(sources_statistics)
        
        # 生成無效轉化警告HTML
        invalid_warning_html = ""
        # 註釋掉原來的警告生成邏輯
        # if invalid_count > 0:
        #     try:
        #         import sys
        #         import os
        #         current_dir = os.path.dirname(os.path.abspath(__file__))
        #         project_root = os.path.join(current_dir, '../../../../')
        #         if project_root not in sys.path:
        #             sys.path.insert(0, project_root)
        #         
        #         from config import get_partner_invalid_warning_config
        #         
        #         # 檢查partner是否配置為顯示invalid warning
        #         show_invalid_warning = get_partner_invalid_warning_config(partner_name)
        #         
        #         # 只有當partner配置為顯示警告且確實有無效轉化時才顯示
        #         if show_invalid_warning:
        #             invalid_warning_html = f"""
        #             <li style="color: #dc3545; font-weight: bold;">
        #                 ⚠️ Invalid Conversion: {invalid_count:,} records (Total Amount: ${invalid_amount:,.2f})
        #             </li>
        #             """
        #     except ImportError:
        #         # 如果無法導入config，使用默認邏輯（向後兼容）
        #         invalid_warning_html = f"""
        #         <li style="color: #dc3545; font-weight: bold;">
        #             ⚠️ Invalid Conversion: {invalid_count:,} records (Total Amount: ${invalid_amount:,.2f})
        #         </li>
        #         """
        
        # 替换模板变量
        body = template
        
        # 调试信息：打印邮件数据
        logger.info(f"🔍 邮件数据调试 - Partner: {partner_name}")
        logger.info(f"   总记录数: {total_records}")
        logger.info(f"   总金额: {total_amount}")
        logger.info(f"   日期范围: {start_date} 至 {end_date}")
        logger.info(f"   有效转化: {email_data.get('pending_approved_count', 0)}")
        logger.info(f"   有效金额: {email_data.get('pending_approved_amount', '$0.00')}")
        logger.info(f"   无效转化: {email_data.get('invalid_rejected_count', 0)}")
        logger.info(f"   无效金额: {email_data.get('invalid_rejected_amount', '$0.00')}")
        
        body = body.replace('{{partner_name}}', partner_name)
        body = body.replace('{{date_range}}', f"{final_start_date} to {final_end_date}")
        body = body.replace('{{start_date}}', final_start_date)
        body = body.replace('{{end_date}}', final_end_date)
        body = body.replace('{{total_records}}', f"{total_records:,}")
        body = body.replace('{{total_amount}}', total_amount)
        body = body.replace('{{main_file}}', main_file)
        body = body.replace('{{sources_list}}', sources_list)
        body = body.replace('{{sources_statistics}}', sources_statistics_html)
        body = body.replace('{{feishu_section}}', feishu_section)
        body = body.replace('{{completion_time}}', completion_time)
        body = body.replace('{{invalid_warning}}', invalid_warning_html)
        
        # 添加新的详细状态统计变量（同步Excel Summary格式）
        body = body.replace('{{total_all_conversions}}', f"{email_data.get('total_all_conversions', 0):,}")
        body = body.replace('{{pending_approved_count}}', f"{email_data.get('pending_approved_count', 0):,}")
        body = body.replace('{{pending_approved_amount}}', email_data.get('pending_approved_amount', '$0.00'))
        body = body.replace('{{invalid_rejected_count}}', f"{email_data.get('invalid_rejected_count', 0):,}")
        body = body.replace('{{invalid_rejected_amount}}', email_data.get('invalid_rejected_amount', '$0.00'))
        
        # 调试信息：检查模板变量替换结果
        logger.info(f"🔍 模板变量替换完成")
        logger.info(f"   模板中是否还有未替换的变量:")
        if '{{date_range}}' in body:
            logger.warning(f"   ⚠️ 发现未替换的变量: {{date_range}}")
        if '{{total_all_conversions}}' in body:
            logger.warning(f"   ⚠️ 发现未替换的变量: {{total_all_conversions}}")
        if '{{pending_approved_count}}' in body:
            logger.warning(f"   ⚠️ 发现未替换的变量: {{pending_approved_count}}")
        if '{{pending_approved_amount}}' in body:
            logger.warning(f"   ⚠️ 发现未替换的变量: {{pending_approved_amount}}")
        if '{{invalid_rejected_count}}' in body:
            logger.warning(f"   ⚠️ 发现未替换的变量: {{invalid_rejected_count}}")
        if '{{invalid_rejected_amount}}' in body:
            logger.warning(f"   ⚠️ 发现未替换的变量: {{invalid_rejected_amount}}")
        
        return body
    
    def _generate_sources_statistics_html(self, sources_statistics):
        """生成Sources统计的HTML（新格式：列表形式）"""
        if not sources_statistics:
            return "<p>No Sources statistics available</p>"
        
        html_parts = ["<ul style='list-style: none; padding: 0; margin: 0;'>"]
        
        for stat in sources_statistics:
            source_name = stat.get('source_name', 'Unknown')
            records = stat.get('records', 0)
            sales_amount = stat.get('sales_amount', '$0.00')
            
            # 确保records是整数类型
            try:
                records = int(records) if records is not None else 0
            except (ValueError, TypeError):
                records = 0
            
            html_parts.append(f"<li style='margin: 8px 0; padding: 8px; background-color: #ffffff; border: 1px solid #e9ecef; border-radius: 4px;'>")
            html_parts.append(f"<strong>- {source_name}:</strong> ")
            html_parts.append(f"Total Conversion: <strong>{records:,}</strong> records, ")
            html_parts.append(f"Total Sale Amount: <span style='color: #28a745; font-weight: bold;'>{sales_amount}</span>")
            html_parts.append("</li>")
        
        html_parts.append("</ul>")
        
        return "".join(html_parts)
    
    def _generate_sources_list(self, sources_statistics):
        """生成Sources列表字符串"""
        if not sources_statistics:
            return "None"
        
        source_names = [stat.get('source_name', 'Unknown') for stat in sources_statistics]
        return ", ".join(source_names)
    
    def _generate_fallback_email_body(self, partner_name, email_data, feishu_info):
        """生成备用的简单邮件正文（当模板加载失败时使用）"""
        report_date = email_data.get('report_date', datetime.now().strftime("%Y-%m-%d"))
        
        # 优先使用从文件名中提取的报表日期
        if 'file_paths' in email_data and email_data['file_paths']:
            extracted_report_date = self._extract_report_date_from_filename(email_data['file_paths'][0])
            if extracted_report_date:
                report_date = extracted_report_date
                print_step("备用邮件日期", f"📅 使用文件名中的报表日期: {report_date}")
        
        completion_time = f"{report_date} {datetime.now().strftime('%H:%M:%S')}"
        
        total_records = email_data.get('total_records', 0)
        total_amount = email_data.get('total_amount', '$0.00')
        start_date = email_data.get('start_date', report_date)
        end_date = email_data.get('end_date', report_date)
        main_file = email_data.get('main_file', f'{partner_name}_ConversionReport_{report_date}.xlsx')
        sources_statistics = email_data.get('sources_statistics', [])
        
        # 生成Sources列表
        sources_list = self._generate_sources_list(sources_statistics)
        
        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <p>Hi {partner_name} Teams,</p>
            <p>{main_file} is attached for your review.</p>
            
            <h3 style="color: #1f4e79;">📊 {partner_name} Report Summary:</h3>
            <ul style="background-color: #f8f9fa; padding: 15px; border-left: 4px solid #007bff;">
                <li><strong>Partner:</strong> {partner_name}</li>
                <li><strong>Date Range:</strong> {start_date} to {end_date}</li>
                <li><strong>Total Conversions (All Status):</strong> {email_data.get('total_all_conversions', 0):,} records</li>
                <li style="color: #28a745; font-weight: bold;"><strong>✅ Total Conversions (Pending/Approved):</strong> {email_data.get('pending_approved_count', 0):,} records</li>
                <li style="color: #28a745; font-weight: bold;"><strong>✅ Total Sale Amount (USD) (Pending/Approved):</strong> {email_data.get('pending_approved_amount', '$0.00')}</li>
                <li style="color: #dc3545; font-weight: bold;"><strong>⚠️ Total Conversions (Invalid/Rejected):</strong> {email_data.get('invalid_rejected_count', 0):,} records</li>
                <li style="color: #dc3545; font-weight: bold;"><strong>⚠️ Total Sale Amount (USD) (Invalid/Rejected):</strong> {email_data.get('invalid_rejected_amount', '$0.00')}</li>
                <li><strong>Sources:</strong> {sources_list}</li>
            </ul>
            <div style="background-color: #f8f9fa; padding: 15px; border-left: 4px solid #007bff; margin-top: 15px;">
                {self._generate_sources_statistics_html(sources_statistics)}
            </div>
            
            <h3 style="color: #1f4e79;">📁 Attachments:</h3>
            <ul><li><strong>{main_file}</strong></li></ul>
        """
        
        # 飞书文件链接
        if feishu_info and self.include_feishu_links:
            if feishu_info.get('success') and feishu_info.get('uploaded_files'):
                body += f'<h3 style="color: #1f4e79;">☁️ Feishu File Links:</h3><ul>'
                for file_info in feishu_info['uploaded_files']:
                    filename = file_info.get('filename', '')
                    
                    # 优先使用飞书返回的URL
                    feishu_url = file_info.get('url')
                    if not feishu_url:
                        # 如果没有url字段，尝试用file_token构造
                        file_token = file_info.get('file_token') or file_info.get('file_id')
                        if file_token:
                            feishu_url = f"https://bytedance.feishu.cn/sheets/{file_token}"
                    
                    if feishu_url:
                        body += f'<li><a href="{feishu_url}" target="_blank" style="color: #007bff; text-decoration: none; font-weight: bold;">{filename}</a></li>'
                    else:
                        body += f'<li>{filename} (Uploaded)</li>'
                body += "</ul>"
        
        body += f"""
            <p style="margin-top: 30px;"><strong>Generated at:</strong> {completion_time}</p>
            <p style="margin-top: 30px; color: #666;">
                Best regards,<br><strong>Reporter-Agent, updated by 2025.07.27</strong>
            </p>
        </body>
        </html>
        """
        
        return body
    
    def _generate_email_body(self, report_data, feishu_upload_result=None):
        """生成邮件正文（兼容性保留）"""
        # 优先使用从文件名中提取的报表日期
        report_date = report_data.get('report_date', datetime.now().strftime("%Y-%m-%d"))
        if 'file_paths' in report_data and report_data['file_paths']:
            extracted_report_date = self._extract_report_date_from_filename(report_data['file_paths'][0])
            if extracted_report_date:
                report_date = extracted_report_date
                print_step("兼容性邮件日期", f"📅 使用文件名中的报表日期: {report_date}")
        
        # 基本信息
        total_records = report_data.get('total_records', 0)
        total_amount = report_data.get('total_amount', '$0.00')
        start_date = report_data.get('start_date', report_date)
        end_date = report_data.get('end_date', report_date)
        completion_time = f"{report_date} {datetime.now().strftime('%H:%M:%S')}"
        
        # 构建邮件正文
        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <p>Hi Partners,</p>
            
            <p>Conversion Report + {start_date} is attached for your review.</p>
            
            <h3 style="color: #1f4e79;">📊 Report Summary:</h3>
            <ul style="background-color: #f8f9fa; padding: 15px; border-left: 4px solid #007bff;">
                <li><strong>Date Range:</strong> {start_date} to {end_date}</li>
                <li><strong>Total Conversions:</strong> {total_records:,} records</li>
                <li><strong>Total Sale Amount:</strong> {total_amount}</li>
            </ul>
            
            <h3 style="color: #1f4e79;">📁 Generated Files:</h3>
            <ul>
        """
        
        # 添加文件信息
        pub_files = report_data.get('pub_files', [])
        main_file = report_data.get('main_file', f'Pub_ConversionReport_{today}.xlsx')
        
        body += f"<li><strong>Main Report:</strong> {os.path.basename(main_file)}</li>"
        
        for pub_file_info in pub_files:
            if isinstance(pub_file_info, dict):
                filename = pub_file_info.get('filename', '')
                records = pub_file_info.get('records', 0)
                amount = pub_file_info.get('amount', '$0.00')
                body += f"<li><strong>{filename}:</strong> ({records:,} records, {amount})</li>"
            else:
                # 如果是文件路径字符串
                filename = os.path.basename(pub_file_info)
                body += f"<li><strong>{filename}</strong></li>"
        
        body += "</ul>"
        
        # 飞书上传状态
        if feishu_upload_result:
            if feishu_upload_result.get('success'):
                feishu_status = f"✅ Success ({feishu_upload_result.get('success_count', 0)} files)"
                if self.include_feishu_links and feishu_upload_result.get('uploaded_files'):
                    body += f"<h3 style='color: #1f4e79;'>☁️ 飞书文件链接:</h3><ul>"
                    for file_info in feishu_upload_result['uploaded_files']:
                        if file_info.get('url'):
                            body += f"<li><a href='{file_info['url']}'>{file_info['filename']}</a></li>"
                    body += "</ul>"
            else:
                feishu_status = f"❌ Failed ({feishu_upload_result.get('failed_count', 0)} files)"
        else:
            feishu_status = "Not executed"
        
        body += f"""
            <h3 style="color: #1f4e79;">☁️ Feishu Upload Status:</h3>
            <p style="background-color: #f8f9fa; padding: 10px; border-radius: 5px;">{feishu_status}</p>
            
            <p style="margin-top: 30px;"><strong>Generated at:</strong> {completion_time}</p>
            
            <p style="margin-top: 30px; color: #666;">
                Best regards,<br>
                <strong>Reporter-Agent, updated by 2025.07.27</strong>
            </p>
        </body>
        </html>
        """
        
        return body
    
    def _compress_file_if_needed(self, file_path, size_threshold_mb=None):
        """如果文件大於閾值，自動壓縮為ZIP格式"""
        try:
            # 使用配置中的閾值，如果沒有提供參數
            if size_threshold_mb is None:
                size_threshold_mb = self.compress_threshold_mb
            
            # 檢查是否啟用自動壓縮
            if not self.auto_compress_attachments:
                print_step("附件處理", f"📎 自動壓縮已禁用，跳過壓縮")
                return file_path, False
            
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            
            if file_size_mb <= size_threshold_mb:
                print_step("附件處理", f"📎 文件 {os.path.basename(file_path)} ({file_size_mb:.1f}MB) 無需壓縮")
                return file_path, False
            
            # 創建臨時ZIP文件
            temp_dir = tempfile.mkdtemp()
            original_filename = os.path.basename(file_path)
            zip_filename = os.path.splitext(original_filename)[0] + '.zip'
            zip_path = os.path.join(temp_dir, zip_filename)
            
            # 壓縮文件
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED, compresslevel=9) as zipf:
                zipf.write(file_path, original_filename)
            
            # 檢查壓縮效果
            compressed_size_mb = os.path.getsize(zip_path) / (1024 * 1024)
            compression_ratio = (1 - compressed_size_mb / file_size_mb) * 100
            
            print_step("附件壓縮", f"✅ 壓縮成功: {original_filename}")
            print_step("附件壓縮", f"   原始大小: {file_size_mb:.1f}MB")
            print_step("附件壓縮", f"   壓縮後: {compressed_size_mb:.1f}MB")
            print_step("附件壓縮", f"   壓縮率: {compression_ratio:.1f}%")
            
            return zip_path, True
            
        except Exception as e:
            print_step("附件壓縮", f"⚠️ 壓縮失敗，使用原文件: {str(e)}")
            return file_path, False

    def _attach_file(self, msg, file_path):
        """添加附件到邮件（支持自動壓縮）"""
        try:
            # 檢查是否需要壓縮
            final_path, is_compressed = self._compress_file_if_needed(file_path)
            
            with open(final_path, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            
            encoders.encode_base64(part)
            filename = os.path.basename(final_path)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= {filename}'
            )
            msg.attach(part)
            
            if is_compressed:
                print_step("附件添加", f"📎 已添加壓縮附件: {filename}")
            else:
                print_step("附件添加", f"📎 已添加附件: {filename}")
            
            # 清理臨時文件
            if is_compressed and final_path != file_path:
                try:
                    os.remove(final_path)
                    os.rmdir(os.path.dirname(final_path))
                except:
                    pass  # 忽略清理錯誤
            
        except Exception as e:
            print_step("附件错误", f"⚠️ 无法添加附件 {file_path}: {str(e)}")
    
    def test_connection(self):
        """测试邮件服务器连接（带超时设置）"""
        print_step("邮件测试", f"正在测试邮件服务器连接... (超时: {self.smtp_timeout}秒)")
        
        if self.password == "your_gmail_app_password_here":
            print_step("配置错误", "❌ 邮件密码未配置")
            return False
        
        try:
            with smtplib.SMTP(self.smtp_server, self.smtp_port, timeout=self.smtp_timeout) as server:
                print_step("邮件测试", "正在建立TLS连接...")
                if self.enable_tls:
                    server.starttls()
                
                print_step("邮件测试", "正在进行身份验证...")
                server.login(self.sender, self.password)
            
            print_step("邮件测试", "✅ 邮件服务器连接成功")
            return True
            
        except (smtplib.SMTPException, socket.timeout, socket.error, ConnectionError, OSError) as e:
            error_type = type(e).__name__
            print_step("邮件测试", f"❌ 邮件服务器连接失败 ({error_type}): {str(e)}")
            return False
        except Exception as e:
            print_step("邮件测试", f"❌ 邮件服务器连接失败 (未知错误): {str(e)}")
            return False

    # 保持向后兼容性的方法别名
    def send_pub_reports(self, pub_summary, feishu_upload_result=None, report_date=None):
        """向后兼容性方法，调用新的send_partner_reports"""
        return self.send_partner_reports(pub_summary, feishu_upload_result, report_date)
    
    # 保持向后兼容性的方法别名
    def _send_single_pub_email(self, pub_name, email_data, file_paths, receivers, feishu_info, report_date=None):
        """向后兼容性方法，调用新的_send_single_partner_email"""
        return self._send_single_partner_email(pub_name, email_data, file_paths, receivers, feishu_info, report_date)
    
    # 保持向后兼容性的方法别名
    def _prepare_pub_email_data(self, pub_name, pub_data, report_date=None):
        """向后兼容性方法，调用新的_prepare_partner_email_data"""
        return self._prepare_partner_email_data(pub_name, pub_data, report_date, None)
    
    # 保持向后兼容性的方法别名
    def _get_pub_feishu_info(self, pub_name, feishu_upload_result):
        """向后兼容性方法，调用新的_get_partner_feishu_info"""
        return self._get_partner_feishu_info(pub_name, feishu_upload_result)
    
    def _create_pub_email_message(self, pub_name, email_data, file_paths, receivers, feishu_info, report_date=None, cc_email=None):
        """向后兼容性方法，调用新的_create_partner_email_message"""
        return self._create_partner_email_message(pub_name, email_data, file_paths, receivers, feishu_info, report_date, cc_email)

    def _is_bytec_partner(self, partner_name):
        """检查是否为ByteC Partner"""
        return config.is_bytec_partner(partner_name)

    def _generate_bytec_email_body(self, partner_name, email_data, feishu_info):
        """生成ByteC专用邮件正文"""
        report_date = email_data.get('report_date', datetime.now().strftime("%Y-%m-%d"))
        
        # 优先使用从文件名中提取的报表日期
        if 'file_paths' in email_data and email_data['file_paths']:
            extracted_report_date = self._extract_report_date_from_filename(email_data['file_paths'][0])
            if extracted_report_date:
                report_date = extracted_report_date
                print_step("ByteC邮件日期", f"📅 使用文件名中的报表日期: {report_date}")
        
        completion_time = f"{report_date} {datetime.now().strftime('%H:%M:%S')}"
        
        # 从email_data获取基本信息
        start_date = email_data.get('start_date', report_date)
        end_date = email_data.get('end_date', report_date)
        main_file = email_data.get('main_file', f'ByteC_ConversionReport_{report_date}.xlsx')
        
        # 加载ByteC专用HTML模板
        template = self._load_html_template('bytec_email_template.html')
        if not template:
            # 如果模板加载失败，使用备用的简单HTML
            return self._generate_fallback_email_body(partner_name, email_data, feishu_info)
        
        # 从Excel文件计算ByteC三个维度的数据
        file_path = email_data.get('file_path')
        if not file_path or not os.path.exists(file_path):
            import traceback
            tb = traceback.format_stack()
            print_step("ByteC邮件", f"⚠️ 文件不存在: {file_path}\n调用栈:\n{''.join(tb)}")
            # 使用默认值
            bytec_data = self._get_default_bytec_data()
        else:
            bytec_data = self._calculate_bytec_summary_from_excel(file_path)
        
        # 准备飞书链接部分
        feishu_section = ""
        if feishu_info and self.include_feishu_links:
            if feishu_info.get('success') and feishu_info.get('uploaded_files'):
                feishu_template = self._load_html_template('feishu_section.html')
                if feishu_template:
                    feishu_links = ""
                    for file_info in feishu_info['uploaded_files']:
                        filename = file_info.get('filename', '')
                        
                        # 优先使用飞书返回的URL
                        feishu_url = file_info.get('url')
                        if not feishu_url:
                            # 如果没有url字段，尝试用file_token构造
                            file_token = file_info.get('file_token') or file_info.get('file_id')
                            if file_token:
                                feishu_url = f"https://bytedance.feishu.cn/sheets/{file_token}"
                        
                        if feishu_url:
                            feishu_links += f'<li><a href="{feishu_url}" target="_blank" style="color: #007bff; text-decoration: none; font-weight: bold;">{filename}</a></li>'
                        else:
                            feishu_links += f'<li>{filename} (已上传)</li>'
                    
                    feishu_section = feishu_template.replace('{{feishu_links}}', feishu_links)
        
        # 替换模板中的占位符
        body = template.replace('{{start_date}}', str(start_date))
        body = body.replace('{{end_date}}', str(end_date))
        body = body.replace('{{main_file}}', main_file)
        body = body.replace('{{completion_time}}', completion_time)
        body = body.replace('{{feishu_section}}', feishu_section)
        
        # ByteC Company Level Summary
        body = body.replace('{{company_total_conversion}}', f"{bytec_data['company']['total_conversion']:,}")
        body = body.replace('{{company_total_sales}}', bytec_data['company']['total_sales'])
        body = body.replace('{{company_total_earning}}', bytec_data['company']['total_earning'])
        body = body.replace('{{company_total_adv_commission}}', bytec_data['company']['total_adv_commission'])
        body = body.replace('{{company_total_pub_commission}}', bytec_data['company']['total_pub_commission'])
        body = body.replace('{{company_total_bytec_commission}}', bytec_data['company']['total_bytec_commission'])
        body = body.replace('{{company_bytec_roi}}', bytec_data['company']['bytec_roi'])
        body = body.replace('{{company_roi_class}}', bytec_data['company']['roi_class'])
        
        # Partner + Source Level Summary
        partner_source_rows = self._generate_partner_source_summary_rows(bytec_data['partner_source'])
        body = body.replace('{{partner_source_summary_rows}}', partner_source_rows)
        
        # Offer Level Summary
        offer_rows = self._generate_offer_summary_rows(bytec_data['offer'])
        body = body.replace('{{offer_summary_rows}}', offer_rows)
        
        return body

    def _calculate_bytec_summary_from_excel(self, file_path):
        """从Excel文件计算ByteC三个维度的汇总数据"""
        try:
            # 读取Excel文件的所有sheet
            excel_file = pd.ExcelFile(file_path)
            all_data = []
            
            print_step("ByteC数据计算", f"📊 开始处理Excel文件: {os.path.basename(file_path)}")
            print_step("ByteC数据计算", f"📋 发现sheets: {excel_file.sheet_names}")
            
            # 優先使用Partner主sheet（避免重複計算）
            partner_main_sheet = None
            for sheet_name in excel_file.sheet_names:
                # Partner主sheet通常沒有下劃線（如DeepLeaper、RAMPUP）
                if '_' not in sheet_name:
                    partner_main_sheet = sheet_name
                    break
            
            if partner_main_sheet:
                # 只讀取Partner主sheet，避免與Source sheets重複
                print_step("ByteC数据计算", f"🎯 使用Partner主sheet: '{partner_main_sheet}'")
                df = pd.read_excel(file_path, sheet_name=partner_main_sheet)
                if not df.empty:
                    all_data.append(df)
                    print_step("ByteC数据计算", f"✅ 成功读取主sheet: {len(df):,} 条记录")
                else:
                    print_step("ByteC数据计算", f"⚠️ 主sheet为空")
            else:
                # 如果沒有主sheet，則合併所有Source sheets
                print_step("ByteC数据计算", f"⚠️ 未找到Partner主sheet，合併所有Source sheets")
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                if not df.empty:
                    all_data.append(df)
                    print_step("ByteC数据计算", f"📋 读取sheet '{sheet_name}': {len(df):,} 条记录")
            
            if not all_data:
                print_step("ByteC数据计算", f"❌ 未找到有效数据")
                return self._get_default_bytec_data()
            
            # 合并数据（通常只有一个Partner主sheet）
            combined_df = pd.concat(all_data, ignore_index=True)
            print_step("ByteC数据计算", f"✅ 合併後總記錄數: {len(combined_df):,}")
            
            # 计算Company Level Summary
            company_summary = self._calculate_company_level_summary(combined_df)
            
            # 计算Partner + Source Level Summary (按优先级排序)
            partner_source_summary = self._calculate_partner_source_summary(combined_df)
            
            # 计算Offer Level Summary (按优先级排序)
            offer_summary = self._calculate_offer_level_summary(combined_df)
            
            return {
                'company': company_summary,
                'partner_source': partner_source_summary,
                'offer': offer_summary
            }
            
        except Exception as e:
            print_step("ByteC数据计算", f"❌ 计算失败: {str(e)}")
            return self._get_default_bytec_data()

    def _calculate_company_level_summary(self, df):
        """计算公司级别汇总"""
        try:
            # 找到转换数量列 - 关键修复！
            conversion_column = None
            for col in ['Conversions', 'conversions', 'conversion', 'Total Conversion']:
                if col in df.columns:
                    conversion_column = col
                    break
            
            # 基本统计：优先使用Conversions列求和，否则用行数
            if conversion_column:
                total_conversion = df[conversion_column].sum()
                print_step("Company汇总", f"🔍 使用'{conversion_column}'列计算转化数: {total_conversion:,}")
            else:
                total_conversion = len(df)
                print_step("Company汇总", f"🔍 使用DataFrame行数计算转化数: {total_conversion:,}")
            
            # 金额计算 - 支持多种列名格式
            sales_amount_column = None
            for col in ['USD Sale Amount', 'sale_amount', 'Sale Amount', 'sales_amount']:
                if col in df.columns:
                    sales_amount_column = col
                    break
            
            total_sales = 0.0
            if sales_amount_column:
                total_sales = df[sales_amount_column].sum()
                print_step("Company汇总", f"🔍 使用'{sales_amount_column}'列计算销售额: ${total_sales:,.2f}")
            
            # Estimated Earning计算
            earning_column = None
            for col in ['Estimated Earning', 'estimated_earning', 'earning']:
                if col in df.columns:
                    earning_column = col
                    break
            
            total_earning = 0.0
            if earning_column:
                total_earning = df[earning_column].sum()
            
            # 佣金计算
            adv_commission_total = 0.0
            pub_commission_total = 0.0
            bytec_commission_total = 0.0
            
            # Adv Commission Rate计算
            adv_commission_column = None
            for col in ['Adv Commission Rate', 'adv_commission_rate']:
                if col in df.columns:
                    adv_commission_column = col
                    break
            
            if adv_commission_column and sales_amount_column:
                # Adv Commission = Sale Amount * Adv Commission Rate
                df_adv = df.copy()
                df_adv[adv_commission_column] = pd.to_numeric(df_adv[adv_commission_column], errors='coerce').fillna(0)
                df_adv[sales_amount_column] = pd.to_numeric(df_adv[sales_amount_column], errors='coerce').fillna(0)
                adv_commission_total = (df_adv[sales_amount_column] * df_adv[adv_commission_column]).sum()
            
            # Pub Commission Rate计算
            pub_commission_column = None
            for col in ['Pub Commission Rate', 'pub_commission_rate']:
                if col in df.columns:
                    pub_commission_column = col
                    break
            
            if pub_commission_column and sales_amount_column:
                # Pub Commission = Sale Amount * Pub Commission Rate
                df_pub = df.copy()
                df_pub[pub_commission_column] = pd.to_numeric(df_pub[pub_commission_column], errors='coerce').fillna(0)
                df_pub[sales_amount_column] = pd.to_numeric(df_pub[sales_amount_column], errors='coerce').fillna(0)
                pub_commission_total = (df_pub[sales_amount_column] * df_pub[pub_commission_column]).sum()
            
            # ByteC Commission = Adv Commission - Pub Commission
            bytec_commission_total = adv_commission_total - pub_commission_total
            
            # ByteC ROI计算 = ByteC Commission / Estimated Earning * 100%
            bytec_roi = 0.0
            roi_class = "amount"  # 默认绿色
            if total_earning > 0:
                bytec_roi = (bytec_commission_total / total_earning) * 100
                if bytec_roi < 0:
                    roi_class = "negative-roi"  # 负值用红色
            
            return {
                'total_conversion': int(total_conversion),  # 确保是整数
                'total_sales': f"${total_sales:,.2f}",
                'total_earning': f"${total_earning:,.2f}",
                'total_adv_commission': f"${adv_commission_total:,.2f}",
                'total_pub_commission': f"${pub_commission_total:,.2f}",
                'total_bytec_commission': f"${bytec_commission_total:,.2f}",
                'bytec_roi': f"{bytec_roi:.2f}%",
                'roi_class': roi_class
            }
            
        except Exception as e:
            print_step("公司汇总计算", f"❌ 失败: {str(e)}")
            return {
                'total_conversion': 0,
                'total_sales': "$0.00",
                'total_earning': "$0.00",
                'total_adv_commission': "$0.00",
                'total_pub_commission': "$0.00",
                'total_bytec_commission': "$0.00",
                'bytec_roi': "0.00%",
                'roi_class': "amount"
            }

    def _calculate_partner_source_summary(self, df):
        """计算Partner + Source维度汇总（按优先级排序）- ByteC Partner+Source汇总增强"""
        try:
            print_step("Partner+Source汇总", "🔍 开始计算Partner+Source汇总...")
            
            # 检查DataFrame是否为空
            if df is None or len(df) == 0:
                print_step("Partner+Source汇总", "⚠️ DataFrame为空，返回空列表")
                return []
            
            print_step("Partner+Source汇总", f"📊 输入数据: {len(df)} 行, {len(df.columns)} 列")
            
            # 找到Partner和Source列
            partner_column = None
            source_column = None
            
            for col in ['Partner', 'partner', 'Partner Name']:
                if col in df.columns:
                    partner_column = col
                    break
            
            for col in ['Source', 'source', 'aff_sub1', 'Source Name']:
                if col in df.columns:
                    source_column = col
                    break
            
            print_step("Partner+Source汇总", f"🔍 找到列: Partner='{partner_column}', Source='{source_column}'")
            
            if not partner_column or not source_column:
                print_step("Partner+Source汇总", f"❌ 缺少必要列: Partner列={partner_column}, Source列={source_column}")
                return []
            
            # 数据清理：移除TOTAL行和空行
            df_clean = df.copy()
            
            # 移除TOTAL行
            if partner_column in df_clean.columns:
                total_mask = (df_clean[partner_column] == 'TOTAL') | (df_clean[partner_column].isnull())
                df_clean = df_clean[~total_mask]
            
            # 移除Source为空的行
            if source_column in df_clean.columns:
                empty_mask = (df_clean[source_column].isnull()) | (df_clean[source_column] == '') | (df_clean[source_column] == 'nan')
                df_clean = df_clean[~empty_mask]
            
            print_step("Partner+Source汇总", f"📊 数据清理后: {len(df_clean)} 行")
            
            if len(df_clean) == 0:
                print_step("Partner+Source汇总", "⚠️ 清理后无有效数据")
                return []
            
            # 找到转换数量列
            conversion_column = None
            for col in ['Conversions', 'conversions', 'conversion', 'Total Conversion']:
                if col in df_clean.columns:
                    conversion_column = col
                    break
            
            # 找到金额列
            sales_amount_column = None
            earning_column = None
            
            for col in ['sale_amount', 'Sale Amount', 'sales_amount']:
                if col in df_clean.columns:
                    sales_amount_column = col
                    break
                    
            for col in ['Estimated Earning', 'estimated_earning', 'earning']:
                if col in df_clean.columns:
                    earning_column = col
                    break
            
            print_step("Partner+Source汇总", f"🔍 数据列: Conversion='{conversion_column}', Sales='{sales_amount_column}', Earning='{earning_column}'")
            
            # 按Partner + Source分组统计
            agg_dict = {}
            if sales_amount_column:
                agg_dict[sales_amount_column] = 'sum'
            if earning_column:
                agg_dict[earning_column] = 'sum'
            if conversion_column:
                agg_dict[conversion_column] = 'sum'
            
            if not agg_dict:
                print_step("Partner+Source汇总", "⚠️ 无可汇总的数据列")
                return []
            
            grouped = df_clean.groupby([partner_column, source_column]).agg(agg_dict).reset_index()
            
            print_step("Partner+Source汇总", f"📊 分组结果: {len(grouped)} 个Partner+Source组合")
            
            # 计算转换数量：优先使用Conversions列，否则用行数
            if conversion_column:
                grouped['conversion_count'] = grouped[conversion_column]
            else:
                # 回退到计算分组行数
                group_counts = df_clean.groupby([partner_column, source_column]).size().reset_index(name='conversion_count')
                grouped = grouped.merge(group_counts, on=[partner_column, source_column])
            
            # 重命名列
            column_mapping = {
                partner_column: partner_column,
                source_column: source_column
            }
            if sales_amount_column:
                column_mapping[sales_amount_column] = 'total_sales'
            if earning_column:
                column_mapping[earning_column] = 'total_earning'
                
            grouped = grouped.rename(columns=column_mapping)
            
            # 创建Partner + Source组合
            grouped['partner_source'] = grouped[partner_column].astype(str) + "+" + grouped[source_column].astype(str)
            
            # 确保数值列存在默认值
            if 'total_sales' not in grouped.columns:
                grouped['total_sales'] = 0.0
            if 'total_earning' not in grouped.columns:
                grouped['total_earning'] = 0.0
            
            # 数据验证
            grouped['total_sales'] = pd.to_numeric(grouped['total_sales'], errors='coerce').fillna(0.0)
            grouped['total_earning'] = pd.to_numeric(grouped['total_earning'], errors='coerce').fillna(0.0)
            grouped['conversion_count'] = pd.to_numeric(grouped['conversion_count'], errors='coerce').fillna(0)
            
            # 按Estimated Earning降序排序（优先级）
            grouped = grouped.sort_values('total_earning', ascending=False)
            
            # 转换为列表格式
            summary_list = []
            for idx, row in grouped.iterrows():
                try:
                    summary_list.append({
                        'partner_source': str(row['partner_source']),
                        'conversion': int(row['conversion_count']),
                        'sales_amount': f"${row['total_sales']:,.2f}",
                        'estimated_earning': f"${row['total_earning']:,.2f}"
                    })
                except Exception as row_error:
                    print_step("Partner+Source汇总", f"⚠️ 行处理错误: {row_error}")
                    continue
            
            print_step("Partner+Source汇总", f"✅ 成功生成 {len(summary_list)} 个Partner+Source汇总")
            
            # 调试输出
            if len(summary_list) > 0:
                print_step("Partner+Source汇总", "📊 汇总详情:")
                for i, item in enumerate(summary_list[:5]):  # 只显示前5个
                    print_step("Partner+Source汇总", f"   {i+1}. {item['partner_source']}: {item['conversion']} conversions, {item['sales_amount']}")
            
            return summary_list
            
        except Exception as e:
            print_step("Partner+Source汇总", f"❌ 失败: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    def _calculate_offer_level_summary(self, df):
        """计算Offer维度汇总（按优先级排序）"""
        try:
            # 找到Offer列
            offer_column = None
            
            for col in ['Offer Name', 'offer_name', 'offer', 'Offer']:
                if col in df.columns:
                    offer_column = col
                    break
            
            if not offer_column:
                return []
            
            # 找到转换数量列 - 关键修复！
            conversion_column = None
            for col in ['Conversions', 'conversions', 'conversion', 'Total Conversion']:
                if col in df.columns:
                    conversion_column = col
                    break
            
            # 找到金额列
            sales_amount_column = None
            earning_column = None
            
            for col in ['USD Sale Amount', 'sale_amount', 'Sale Amount', 'sales_amount']:
                if col in df.columns:
                    sales_amount_column = col
                    break
                    
            for col in ['Estimated Earning', 'estimated_earning', 'earning']:
                if col in df.columns:
                    earning_column = col
                    break
            
            # 过滤掉TOTAL行（这是Excel中的汇总行，不是真实的Offer）
            df_filtered = df[df[offer_column] != 'TOTAL'].copy()
            
            # 按Offer分组统计
            agg_dict = {}
            if sales_amount_column:
                agg_dict[sales_amount_column] = 'sum'
            if earning_column:
                agg_dict[earning_column] = 'sum'
            # 关键修复：如果有Conversions列，则求和；否则用行数
            if conversion_column:
                agg_dict[conversion_column] = 'sum'
            
            grouped = df_filtered.groupby(offer_column).agg(agg_dict).reset_index()
            
            # 计算转换数量：优先使用Conversions列，否则用行数
            if conversion_column:
                # 直接使用Conversions列的汇总值
                grouped['conversion_count'] = grouped[conversion_column]
            else:
                # 回退到计算分组行数（兼容性）
                group_counts = df_filtered.groupby(offer_column).size().reset_index(name='conversion_count')
                grouped = grouped.merge(group_counts, on=offer_column)
            
            # 重命名列
            column_mapping = {
                offer_column: offer_column
            }
            if sales_amount_column:
                column_mapping[sales_amount_column] = 'total_sales'
            if earning_column:
                column_mapping[earning_column] = 'total_earning'
                
            grouped = grouped.rename(columns=column_mapping)
            
            # 确保missing列存在默认值
            if 'total_sales' not in grouped.columns:
                grouped['total_sales'] = 0.0
            if 'total_earning' not in grouped.columns:
                grouped['total_earning'] = 0.0
            
            # 按Estimated Earning降序排序（优先级）
            grouped = grouped.sort_values('total_earning', ascending=False)
            
            # 转换为列表格式
            summary_list = []
            for idx, row in grouped.iterrows():
                summary_list.append({
                    'offer_name': row[offer_column],
                    'conversion': int(row['conversion_count']),
                    'sales_amount': f"${row['total_sales']:,.2f}",
                    'estimated_earning': f"${row['total_earning']:,.2f}"
                })
            
            return summary_list
            
        except Exception as e:
            print_step("Offer汇总", f"❌ 失败: {str(e)}")
            return []

    def _generate_partner_source_summary_rows(self, partner_source_data):
        """生成Partner + Source汇总表格行"""
        if not partner_source_data:
            return "<tr><td colspan='5'>暂无数据</td></tr>"
        
        rows = []
        for idx, item in enumerate(partner_source_data, 1):
            row = f"""
            <tr>
                <td><strong>{idx}</strong></td>
                <td><strong>{item['partner_source']}</strong></td>
                <td>{item['conversion']:,}</td>
                <td><span class="amount">{item['sales_amount']}</span></td>
                <td><span class="amount">{item['estimated_earning']}</span></td>
            </tr>
            """
            rows.append(row)
        
        return "".join(rows)

    def _generate_offer_summary_rows(self, offer_data):
        """生成Offer汇总表格行"""
        if not offer_data:
            return "<tr><td colspan='5'>暂无数据</td></tr>"
        
        rows = []
        for idx, item in enumerate(offer_data, 1):
            row = f"""
            <tr>
                <td><strong>{idx}</strong></td>
                <td><strong>{item['offer_name']}</strong></td>
                <td>{item['conversion']:,}</td>
                <td><span class="amount">{item['sales_amount']}</span></td>
                <td><span class="amount">{item['estimated_earning']}</span></td>
            </tr>
            """
            rows.append(row)
        
        return "".join(rows)

    def _get_default_bytec_data(self):
        """获取默认的ByteC数据（当无法读取Excel时使用）"""
        return {
            'company': {
                'total_conversion': 0,
                'total_sales': "$0.00",
                'total_earning': "$0.00",
                'total_adv_commission': "$0.00",
                'total_pub_commission': "$0.00",
                'total_bytec_commission': "$0.00",
                'bytec_roi': "0.00%",
                'roi_class': "amount"
            },
            'partner_source': [],
            'offer': []
        }

    def _should_fallback_to_cloud_link(self, msg_size_mb, failed_attempts):
        """檢查是否應該降級到雲端鏈接模式"""
        # 檢查是否啟用智能降級策略
        if not self.smart_fallback_enabled:
            return False
        
        # 條件1：文件過大
        if msg_size_mb > self.fallback_size_threshold_mb:
            return True
        
        # 條件2：重試失敗次數超過閾值
        if failed_attempts >= self.fallback_retry_threshold:
            return True
            
        return False
    
    def _fallback_to_cloud_link_mode(self, original_msg, recipients, operation_name):
        """降級到雲端鏈接模式：發送無附件郵件，提供飛書下載鏈接"""
        try:
            print_step(f"{operation_name}雲端模式", "☁️ 切換到雲端鏈接模式，移除附件...")
            
            # 創建無附件版本的郵件
            fallback_msg = MIMEMultipart()
            fallback_msg['From'] = original_msg['From']
            fallback_msg['To'] = original_msg['To']
            fallback_msg['Cc'] = original_msg.get('Cc', '')
            fallback_msg['Subject'] = original_msg['Subject']
            
            # 獲取原始郵件內容
            original_body = None
            for part in original_msg.walk():
                if part.get_content_type() == "text/html":
                    original_body = part.get_payload(decode=True).decode('utf-8')
                    break
            
            if original_body:
                # 在郵件末尾添加雲端下載說明
                cloud_notice = """
                <div style='margin-top: 30px; padding: 20px; background-color: #e8f4fd; border: 1px solid #bee5eb; border-radius: 8px;'>
                    <h3 style='color: #0c5460; margin-top: 0;'>📁 文件下載說明</h3>
                    <p style='color: #0c5460; margin-bottom: 10px;'>由於附件較大，已上傳至飛書雲端存儲，請通過以下方式獲取：</p>
                    <ol style='color: #0c5460; margin-bottom: 0;'>
                        <li>聯繫系統管理員獲取飛書文件鏈接</li>
                        <li>或透過飛書工作台查看最新上傳的Excel報告</li>
                    </ol>
                    <p style='color: #6c757d; font-size: 0.9em; margin-top: 15px; margin-bottom: 0;'>
                        💡 提示：雲端文件支援線上預覽和下載，安全便捷
                    </p>
                </div>
                """
                
                # 將雲端說明插入到郵件內容中
                enhanced_body = original_body.replace('</body>', cloud_notice + '</body>')
                fallback_msg.attach(MIMEText(enhanced_body, 'html'))
            else:
                # 如果無法獲取原始內容，創建簡單說明
                simple_body = """
                <html><body>
                <h2>📊 轉換報告</h2>
                <p>由於技術原因，附件已上傳至飛書雲端存儲。</p>
                <p>請聯繫系統管理員獲取文件鏈接。</p>
                </body></html>
                """
                fallback_msg.attach(MIMEText(simple_body, 'html'))
            
            print_step(f"{operation_name}雲端模式", "📧 準備發送無附件郵件...")
            
            # 使用簡化的SMTP發送（無重試，因為無附件應該很快）
            try:
                with smtplib.SMTP(self.smtp_server, self.smtp_port, timeout=60) as server:
                    if self.enable_tls:
                        server.starttls()
                    server.login(self.sender, self.password)
                    server.send_message(fallback_msg)
                
                print_step(f"{operation_name}雲端模式", "✅ 雲端鏈接模式郵件發送成功")
                return {'success': True, 'attempts': 1, 'fallback_mode': 'cloud_link'}
                
            except Exception as fallback_error:
                print_step(f"{operation_name}雲端模式", f"❌ 雲端模式也失敗: {str(fallback_error)}")
                return {'success': False, 'error': str(fallback_error), 'fallback_mode': 'cloud_link'}
                
        except Exception as e:
            print_step(f"{operation_name}雲端模式", f"❌ 降級處理失敗: {str(e)}")
            return {'success': False, 'error': str(e)}

    def _send_email_with_retry(self, msg, recipients, operation_name="邮件发送"):
        """
        增强版重试机制的稳定邮件发送方法
        
        Args:
            msg: 邮件消息对象
            recipients: 收件人列表
            operation_name: 操作名称（用于日志）
        
        Returns:
            dict: 发送结果
        """
        last_error = None
        delay = self.retry_delay
        
        # 🚀 階段1優化：動態超時調整（基於文件大小智能調整）
        msg_size = len(msg.as_string())
        msg_size_mb = msg_size / (1024 * 1024)
        
        # 根據文件大小智能調整超時時間（使用配置選項）
        if self.dynamic_timeout_enabled:
            if msg_size_mb > 15:  # 大文件 (>15MB)
                dynamic_timeout = self.large_file_timeout
                file_category = "大文件"
            elif msg_size_mb > 5:  # 中等文件 (5-15MB)
                dynamic_timeout = self.medium_file_timeout
                file_category = "中等文件"
            else:  # 小文件 (<5MB)
                dynamic_timeout = self.small_file_timeout
                file_category = "小文件"
        else:
            # 如果禁用動態超時，使用默認超時
            dynamic_timeout = self.smtp_timeout
            file_category = "文件"
        
        print_step(f"{operation_name}{file_category}", f"📎 檢測到{file_category} ({msg_size_mb:.1f}MB)，超時調整為 {dynamic_timeout}秒")
        
        # 智能重试策略：根据错误类型调整重试策略
        retryable_errors = (
            smtplib.SMTPException, 
            smtplib.SMTPServerDisconnected,  # 添加服务器断开连接错误
            socket.timeout, 
            socket.error, 
            ConnectionError, 
            OSError,
            ConnectionResetError,
            BrokenPipeError
        )
        
        for attempt in range(self.max_retries + 1):  # +1 因为第一次不算重试
            try:
                print_step(f"{operation_name}尝试", f"🔄 第 {attempt + 1} 次尝试 (超时设置: {dynamic_timeout}秒)")
                
                # 创建SMTP连接，设置动态超时
                with smtplib.SMTP(self.smtp_server, self.smtp_port, timeout=dynamic_timeout) as server:
                                          # 设置调试模式（仅在开发时启用）
                      # server.set_debuglevel(1)
                    
                    if self.enable_tls:
                        print_step(f"{operation_name}TLS", "🔒 启动TLS加密连接...")
                        server.starttls()
                    
                    print_step(f"{operation_name}认证", "🔐 正在登录SMTP服务器...")
                    server.login(self.sender, self.password)
                    
                    print_step(f"{operation_name}发送", f"📤 正在发送邮件给 {len(recipients)} 个收件人...")
                    # 使用send_message方法，更可靠和简洁（自动从邮件头读取收件人）
                    server.send_message(msg)
                
                # 发送成功
                print_step(f"{operation_name}成功", f"✅ 邮件发送成功 (第 {attempt + 1} 次尝试)")
                return {'success': True, 'attempts': attempt + 1}
                
            except retryable_errors as e:
                last_error = e
                error_type = type(e).__name__
                error_msg = str(e)
                
                # 智能错误分析
                if "Connection unexpectedly closed" in error_msg:
                    error_desc = "连接意外关闭"
                elif "timeout" in error_msg.lower():
                    error_desc = "连接超时"
                elif "connection" in error_msg.lower():
                    error_desc = "连接错误"
                else:
                    error_desc = "SMTP错误"
                
                print_step(f"{operation_name}错误", f"❌ 第 {attempt + 1} 次尝试失败 ({error_type}): {error_desc}")
                
                # 如果不是最后一次尝试，则等待后重试
                if attempt < self.max_retries:
                    # 智能延迟：根据错误类型调整延迟时间
                    if "timeout" in error_msg.lower():
                        smart_delay = min(delay * 1.5, self.max_retry_delay)  # 超时错误增加延迟
                    elif "Connection unexpectedly closed" in error_msg:
                        smart_delay = min(delay * 2, self.max_retry_delay)  # 连接关闭错误增加更多延迟
                    else:
                        smart_delay = min(delay, self.max_retry_delay)
                    
                    print_step(f"{operation_name}重试", f"⏳ 等待 {smart_delay:.1f} 秒后重试...")
                    time.sleep(smart_delay)
                    delay = min(delay * self.retry_backoff, self.max_retry_delay)  # 指数退避但不超过最大值
                else:
                    # 🛡️ 階段2優化：智能降級檢查
                    if self._should_fallback_to_cloud_link(msg_size_mb, attempt + 1):
                        print_step(f"{operation_name}降級", f"🔄 觸發智能降級條件，嘗試雲端鏈接模式...")
                        return self._fallback_to_cloud_link_mode(msg, recipients, operation_name)
                    else:
                        print_step(f"{operation_name}失败", f"❌ 所有重試均失敗，放弃发送")
            
            except Exception as e:
                # 其他未预期的错误，不重试
                last_error = e
                print_step(f"{operation_name}异常", f"❌ 发生未预期错误: {str(e)}")
                break
        
        # 所有尝试都失败了
        return {
            'success': False, 
            'error': f"邮件发送失败 (尝试 {self.max_retries + 1} 次): {str(last_error)}",
            'attempts': self.max_retries + 1
        }

    def _is_partner_email_enabled(self, partner_name: str) -> bool:
        """檢查Partner郵件是否啟用（大小寫不敏感）"""
        # 先嘗試直接匹配
        if partner_name in self.partner_email_enabled:
            return self.partner_email_enabled[partner_name]
        
        # 如果直接匹配失敗，嘗試大小寫不敏感匹配
        partner_name_lower = partner_name.lower()
        for config_key, enabled in self.partner_email_enabled.items():
            if config_key.lower() == partner_name_lower:
                return enabled
        
        # 如果都沒有匹配，返回 False
        return False
    
    def _get_partner_recipients(self, partner_name: str, self_email: bool = False) -> List[str]:
        """獲取Partner的收件人列表（大小寫不敏感）"""
        if self_email:
            return [self.sender]
        
        # 使用 config.get_partner_email_config 獲取最新配置
        try:
            email_config = config.get_partner_email_config(partner_name)
            recipients = email_config.get('recipients', [])
            if recipients:
                return recipients
        except Exception:
            pass
        
        # 備用方案：嘗試大小寫不敏感匹配
        partner_name_lower = partner_name.lower()
        for config_key, recipients in self.partner_email_mapping.items():
            if config_key.lower() == partner_name_lower:
                return recipients
        
        # 最後備用方案
        return self.default_receivers
    
    def _generate_email_subject(self, partner_name: str, report_date: str, file_path: str = None) -> str:
        """生成郵件主題"""
        if file_path and os.path.exists(file_path):
            # 使用檔名作為郵件主題
            filename = os.path.basename(file_path)
            return filename
        else:
            # 備用方案
            return f"{partner_name} Conversion Report - {report_date}"
    
    def _test_smtp_connection(self) -> bool:
        """測試SMTP連接"""
        try:
            with smtplib.SMTP(self.smtp_server, self.smtp_port, timeout=self.smtp_timeout) as server:
                if self.enable_tls:
                    server.starttls()
                server.login(self.sender, self.password)
                return True
        except Exception as e:
            print_step("SMTP測試", f"❌ SMTP連接失敗: {e}")
            return False
    
    def _send_single_email(self, recipients: List[str], subject: str, body: str, 
                          attachments: List[str] = None, is_html: bool = True) -> bool:
        """發送單封郵件"""
        try:
            msg = MIMEMultipart()
            msg['From'] = self.sender
            msg['To'] = ', '.join(recipients)
            msg['Subject'] = subject
            
            # 添加抄送
            if self.auto_cc_email:
                msg['Cc'] = self.auto_cc_email
                recipients.append(self.auto_cc_email)
            
            # 添加郵件正文
            msg.attach(MIMEText(body, 'html' if is_html else 'plain'))
            
            # 添加附件
            if attachments:
                for file_path in attachments:
                    if os.path.exists(file_path):
                        with open(file_path, "rb") as attachment:
                            part = MIMEBase('application', 'octet-stream')
                            part.set_payload(attachment.read())
                        
                        encoders.encode_base64(part)
                        part.add_header(
                            'Content-Disposition',
                            f'attachment; filename= {os.path.basename(file_path)}'
                        )
                        msg.attach(part)
            
            # 發送郵件
            with smtplib.SMTP(self.smtp_server, self.smtp_port, timeout=self.smtp_timeout) as server:
                if self.enable_tls:
                    server.starttls()
                server.login(self.sender, self.password)
                server.send_message(msg)
            
            return True
            
        except Exception as e:
            print_step("郵件發送", f"❌ 發送郵件失敗: {e}")
            return False

    def _calculate_total_sales_amount_from_excel(self, file_path):
        """🚀 優化版本：計算Excel總金額，避免重複讀取"""
        import time
        start_time = time.time()
        
        try:
            if not file_path or not os.path.exists(file_path):
                import traceback
                tb = traceback.format_stack()
                print_step("金额计算", f"⚠️ 文件不存在: {file_path}\n调用栈:\n{''.join(tb)}")
                return '$0.00'
            
            print_step("Excel處理", f"🚀 開始處理 {os.path.basename(file_path)}...")
            
            # 🚀 性能優化1：一次性讀取所有sheets
            read_start = time.time()
            all_sheets_data = pd.read_excel(file_path, sheet_name=None)
            read_time = time.time() - read_start
            print_step("讀取優化", f"✅ 一次讀取完成 {len(all_sheets_data)} sheets，耗時 {read_time:.2f}秒")
            
            total_amount = 0.0
            sheet_details = []
            
            # 獲取partner名稱
            partner_name = os.path.basename(file_path).split('_')[0]
            
            # 🚀 性能優化2：智能處理邏輯，避免重複處理
            process_start = time.time()
            
            for sheet_name, df in all_sheets_data.items():
                try:
                    # 判断是否为Partner主sheet
                    is_partner_sheet = (sheet_name == partner_name)
                    
                    # 🚀 優化：智能檢測列名，只嘗試常見的skiprows值
                    if 'USD Sale Amount' not in df.columns:
                        for skip_rows in [1, 2, 3]:  # 只嘗試最常見的3個值
                            try:
                                temp_df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=skip_rows, nrows=1)
                                if 'USD Sale Amount' in temp_df.columns:
                                    df = all_sheets_data[sheet_name] = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=skip_rows)
                                    break
                            except:
                                continue
                    
                    # 计算该sheet的金额
                    if 'USD Sale Amount' in df.columns:
                        sheet_amount = df['USD Sale Amount'].fillna(0).sum()
                        total_amount += sheet_amount
                        sheet_details.append(f"  📊 {sheet_name}: ${sheet_amount:,.2f}")
                    
                except Exception as e:
                    print_step("金额计算", f"⚠️ 处理sheet {sheet_name} 失败: {e}")
                    continue
            
            process_time = time.time() - process_start
            total_time = time.time() - start_time
            
            formatted_amount = f"${total_amount:,.2f}"
            print_step("金额计算", f"💰 {os.path.basename(file_path)} 总销售额: {formatted_amount}")
            print_step("金额详情", f"各Sheet明细:\n" + "\n".join(sheet_details))
            print_step("性能統計", f"✅ 優化完成 - 處理: {process_time:.2f}s, 總計: {total_time:.2f}s")
            
            return formatted_amount
                
        except Exception as e:
            print_step("金额计算", f"❌ 计算金额失败 {os.path.basename(file_path)}: {str(e)}")
            return '$0.00'

# 便捷函数
def send_report_email(report_data, file_paths=None, feishu_upload_result=None):
    """
    便捷的邮件发送函数
    
    Args:
        report_data: 报告数据字典
        file_paths: Excel文件路径列表
        feishu_upload_result: 飞书上传结果
    
    Returns:
        dict: 发送结果
    """
    sender = EmailSender()
    return sender.send_report_email(report_data, file_paths, feishu_upload_result)

def test_email_connection():
    """
    测试邮件连接的便捷函数
    
    Returns:
        bool: 连接是否成功
    """
    sender = EmailSender()
    return sender.test_connection() 