"""
優化的 Excel 讀取器模組
提供智能數據檢測和高效讀取功能
"""

import pandas as pd
import openpyxl
from typing import Dict, List, Tuple, Optional, Union
import logging
import os

logger = logging.getLogger(__name__)


class ExcelReader:
    """優化的 Excel 讀取器"""
    
    def __init__(self):
        """初始化 Excel 讀取器"""
        # 常見的數據列名模式
        self._common_data_columns = [
            'USD Sale Amount', 'sale_amount', 'Sale Amount', 'sales_amount',
            'Status', 'status', 'conversion_status',
            'Partner', 'partner', 'Partner Name',
            'Source', 'source', 'aff_sub1', 'Source Name',
            'Order ID', 'order_id', 'Conversion ID', 'conversion_id'
        ]
        
        # 統計信息
        self._stats = {
            'files_read': 0,
            'sheets_read': 0,
            'smart_detections': 0,
            'fallback_attempts': 0
        }
    
    def find_data_start_row(self, file_path: str, sheet_name: str, 
                           target_columns: List[str] = None) -> Tuple[pd.DataFrame, int]:
        """
        智能檢測數據開始行
        
        Args:
            file_path: Excel 文件路徑
            sheet_name: Sheet 名稱
            target_columns: 目標列名列表，如果 None 則使用常見列名
            
        Returns:
            (DataFrame, skip_rows): 數據 DataFrame 和跳過的行數
        """
        if target_columns is None:
            target_columns = self._common_data_columns
        
        logger.debug(f"智能檢測數據開始行: {sheet_name}")
        
        try:
            # 快速檢測：先讀取前 10 行來檢查
            peek_df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=10)
            
            # 檢查第一行是否包含目標列
            if self._has_target_columns(peek_df.columns, target_columns):
                self._stats['smart_detections'] += 1
                # 直接讀取完整數據
                full_df = pd.read_excel(file_path, sheet_name=sheet_name)
                logger.debug(f"智能檢測成功: {sheet_name} - 數據從第1行開始")
                return full_df, 0
            
            # 如果第一行不是數據，嘗試檢測實際數據開始行
            data_start_row = self._detect_data_start_row(file_path, sheet_name, target_columns)
            
            if data_start_row > 0:
                self._stats['smart_detections'] += 1
                full_df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=data_start_row)
                logger.debug(f"智能檢測成功: {sheet_name} - 數據從第{data_start_row + 1}行開始")
                return full_df, data_start_row
            
            # 如果智能檢測失敗，回退到原始方法
            return self._fallback_detection(file_path, sheet_name, target_columns)
            
        except Exception as e:
            logger.warning(f"智能檢測失敗 {sheet_name}: {str(e)}，使用回退檢測")
            return self._fallback_detection(file_path, sheet_name, target_columns)
    
    def _has_target_columns(self, columns: pd.Index, target_columns: List[str]) -> bool:
        """檢查是否包含目標列"""
        columns_set = set(str(col).strip() for col in columns)
        target_set = set(target_columns)
        
        # 檢查是否有任何目標列匹配
        intersection = columns_set.intersection(target_set)
        return len(intersection) > 0
    
    def _detect_data_start_row(self, file_path: str, sheet_name: str, 
                              target_columns: List[str]) -> int:
        """
        檢測數據開始行（高效版本）
        
        Args:
            file_path: Excel 文件路徑
            sheet_name: Sheet 名稱
            target_columns: 目標列名列表
            
        Returns:
            數據開始行號（0-based），如果檢測失敗返回 -1
        """
        try:
            # 使用 openpyxl 快速掃描前 10 行
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb[sheet_name]
            
            # 檢查前 10 行
            for row_idx in range(1, min(11, ws.max_row + 1)):
                row_values = []
                for col_idx in range(1, min(21, ws.max_column + 1)):  # 檢查前 20 列
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        row_values.append(str(cell.value).strip())
                
                # 檢查這一行是否包含目標列名
                if self._has_target_columns(pd.Index(row_values), target_columns):
                    wb.close()
                    return row_idx - 1  # 轉換為 0-based
            
            wb.close()
            return -1  # 未檢測到
            
        except Exception as e:
            logger.warning(f"openpyxl 檢測失敗: {str(e)}")
            return -1
    
    def _fallback_detection(self, file_path: str, sheet_name: str, 
                           target_columns: List[str]) -> Tuple[pd.DataFrame, int]:
        """
        回退檢測方法（減少嘗試次數）
        
        Args:
            file_path: Excel 文件路徑
            sheet_name: Sheet 名稱
            target_columns: 目標列名列表
            
        Returns:
            (DataFrame, skip_rows): 數據 DataFrame 和跳過的行數
        """
        self._stats['fallback_attempts'] += 1
        logger.debug(f"使用回退檢測: {sheet_name}")
        
        # 減少嘗試次數從 15 次到 5 次
        for skip_rows in range(1, 6):
            try:
                test_df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=skip_rows)
                if self._has_target_columns(test_df.columns, target_columns):
                    logger.debug(f"回退檢測成功: {sheet_name} - 跳過 {skip_rows} 行")
                    return test_df, skip_rows
            except Exception:
                continue
        
        # 如果所有嘗試都失敗，返回原始數據
        logger.warning(f"所有檢測方法失敗: {sheet_name}，返回原始數據")
        try:
            original_df = pd.read_excel(file_path, sheet_name=sheet_name)
            return original_df, 0
        except Exception as e:
            logger.error(f"讀取原始數據失敗: {sheet_name} - {str(e)}")
            return pd.DataFrame(), 0
    
    def read_sheet_with_detection(self, file_path: str, sheet_name: str, 
                                 target_columns: List[str] = None) -> pd.DataFrame:
        """
        使用智能檢測讀取 Sheet
        
        Args:
            file_path: Excel 文件路徑
            sheet_name: Sheet 名稱
            target_columns: 目標列名列表
            
        Returns:
            讀取的 DataFrame
        """
        df, skip_rows = self.find_data_start_row(file_path, sheet_name, target_columns)
        self._stats['sheets_read'] += 1
        return df
    
    def read_excel_optimized(self, file_path: str, 
                            target_columns: List[str] = None) -> Dict[str, pd.DataFrame]:
        """
        優化讀取整個 Excel 文件
        
        Args:
            file_path: Excel 文件路徑
            target_columns: 目標列名列表
            
        Returns:
            字典，key 為 sheet 名稱，value 為 DataFrame
        """
        if not os.path.exists(file_path):
            import traceback
            tb = traceback.format_stack()
            logger.error(f"文件不存在: {file_path}\n调用栈:\n{''.join(tb)}")
            return {}
        
        self._stats['files_read'] += 1
        sheets_data = {}
        
        try:
            # 獲取所有 sheet 名稱
            with pd.ExcelFile(file_path) as excel_file:
                sheet_names = excel_file.sheet_names
            
            logger.info(f"開始讀取 Excel 文件: {os.path.basename(file_path)}, {len(sheet_names)} 個 sheets")
            
            # 為每個 sheet 進行智能讀取
            for sheet_name in sheet_names:
                try:
                    df = self.read_sheet_with_detection(file_path, sheet_name, target_columns)
                    sheets_data[sheet_name] = df
                    logger.debug(f"成功讀取 Sheet '{sheet_name}': {len(df)} 行")
                except Exception as e:
                    logger.warning(f"讀取 Sheet '{sheet_name}' 失敗: {str(e)}")
                    sheets_data[sheet_name] = pd.DataFrame()
            
            logger.info(f"Excel 文件讀取完成: {os.path.basename(file_path)}")
            
        except Exception as e:
            logger.error(f"讀取 Excel 文件失敗: {os.path.basename(file_path)} - {str(e)}")
        
        return sheets_data
    
    def get_stats(self) -> Dict[str, int]:
        """獲取讀取統計信息"""
        total_detections = self._stats['smart_detections'] + self._stats['fallback_attempts']
        smart_rate = (self._stats['smart_detections'] / total_detections * 100) if total_detections > 0 else 0
        
        return {
            'files_read': self._stats['files_read'],
            'sheets_read': self._stats['sheets_read'],
            'smart_detections': self._stats['smart_detections'],
            'fallback_attempts': self._stats['fallback_attempts'],
            'smart_detection_rate_percent': round(smart_rate, 2)
        }
    
    def reset_stats(self):
        """重置統計信息"""
        self._stats = {
            'files_read': 0,
            'sheets_read': 0,
            'smart_detections': 0,
            'fallback_attempts': 0
        }


# 全局讀取器實例
_global_excel_reader = None


def get_excel_reader() -> ExcelReader:
    """獲取全局 Excel 讀取器實例"""
    global _global_excel_reader
    if _global_excel_reader is None:
        _global_excel_reader = ExcelReader()
    return _global_excel_reader


def read_excel_optimized(file_path: str, target_columns: List[str] = None) -> Dict[str, pd.DataFrame]:
    """
    便捷函數：優化讀取 Excel 文件
    
    Args:
        file_path: Excel 文件路徑
        target_columns: 目標列名列表
        
    Returns:
        字典，key 為 sheet 名稱，value 為 DataFrame
    """
    reader = get_excel_reader()
    return reader.read_excel_optimized(file_path, target_columns)


def find_data_start_row(file_path: str, sheet_name: str, 
                       target_columns: List[str] = None) -> Tuple[pd.DataFrame, int]:
    """
    便捷函數：智能檢測數據開始行
    
    Args:
        file_path: Excel 文件路徑
        sheet_name: Sheet 名稱
        target_columns: 目標列名列表
        
    Returns:
        (DataFrame, skip_rows): 數據 DataFrame 和跳過的行數
    """
    reader = get_excel_reader()
    return reader.find_data_start_row(file_path, sheet_name, target_columns)


def reset_excel_reader():
    """重置全局 Excel 讀取器（主要用於測試）"""
    global _global_excel_reader
    _global_excel_reader = None 